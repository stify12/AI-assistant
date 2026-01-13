from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "批量对比模板"

# 样式定义
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

# 表头
headers = ['题号', '基准答案', '批改结果1', '批改结果2', '批改结果3']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center_align

# 示例数据
example_data = [
    ['1', 'A', 'A', 'A', 'B'],
    ['2', 'B', 'B', 'C', 'B'],
    ['3', 'C', 'C', 'C', 'C'],
    ['4', 'D', 'D', 'D', 'A'],
    ['5', 'A', 'B', 'A', 'A'],
    ['6', '√', '√', '×', '√'],
    ['7', '3.14', '3.14', '3.14', '3.15'],
    ['8', '北京', '北京', '北京', '上海'],
]

for row_idx, row_data in enumerate(example_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = center_align

# 设置列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15

# 添加说明sheet
ws2 = wb.create_sheet(title="使用说明")
instructions = [
    "批量对比模板使用说明",
    "",
    "格式要求：",
    "1. 第一列（A列）：题号 - 可以是数字、字母或其他标识",
    "2. 第二列（B列）：基准答案 - 标准正确答案",
    "3. 后续列（C列起）：各次批改结果 - 列名会作为标签显示",
    "",
    "支持的答案格式：",
    "- 选择题：A, B, C, D 等",
    "- 判断题：√, ×, 对, 错, T, F 等",
    "- 填空题：任意文本",
    "- 计算题：数字或表达式",
    "",
    "JSON数组填入示例：",
    '[{"index":"1","answer":"A"},{"index":"2","answer":"B"}]',
    "",
    "注意事项：",
    "- 答案对比采用精确匹配",
    "- 可以添加更多批改结果列",
    "- 确保每行都有题号和基准答案",
]

for row_idx, text in enumerate(instructions, 1):
    cell = ws2.cell(row=row_idx, column=1, value=text)
    if row_idx == 1:
        cell.font = Font(bold=True, size=14)
    elif text.startswith('['):
        cell.font = Font(name='Consolas', size=10)

ws2.column_dimensions['A'].width = 60

# 保存
wb.save('batch_compare_template.xlsx')
print("模板已创建: batch_compare_template.xlsx")
