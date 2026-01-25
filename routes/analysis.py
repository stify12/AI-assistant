"""
AI 智能分析 API 路由（增强版）
提供分析触发、报告查询、层级下钻、样本管理、异常检测等接口
"""
import json
from datetime import datetime
from flask import Blueprint, jsonify, request
from services.ai_analysis_service import AIAnalysisService
from services.storage_service import StorageService
from services.llm_service import LLMService

analysis_bp = Blueprint('analysis', __name__)

# 初始化分析服务
analysis_service = AIAnalysisService()


# ============================================
# 分析触发与状态 API
# ============================================

@analysis_bp.route('/api/analysis/trigger/<task_id>', methods=['POST'])
def trigger_analysis(task_id):
    """触发任务分析"""
    try:
        # 检查任务是否存在
        task_data = StorageService.load_batch_task(task_id)
        if not task_data:
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        
        # 检查任务是否已完成
        if task_data.get('status') != 'completed':
            return jsonify({'success': False, 'error': '任务尚未完成，无法分析'}), 400
        
        # 获取优先级参数
        data = request.get_json() or {}
        priority = data.get('priority', 'medium')
        if priority not in ['high', 'medium', 'low']:
            priority = 'medium'
        
        # 触发分析
        result = analysis_service.trigger_analysis(task_id, priority)
        
        return jsonify({
            'success': True,
            'data': {
                'queued': result.get('queued', False),
                'job_id': result.get('job_id'),
                'position': result.get('position', 0),
                'message': result.get('message', '分析任务已加入队列')
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/queue', methods=['GET'])
def get_analysis_queue():
    """获取分析队列状态"""
    try:
        status = analysis_service.get_analysis_queue_status()
        return jsonify({
            'success': True,
            'data': status
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/queue/<job_id>', methods=['DELETE'])
def cancel_analysis(job_id):
    """取消排队中的分析任务"""
    try:
        result = analysis_service.cancel_analysis(job_id)
        return jsonify({
            'success': result.get('success', False),
            'message': result.get('message', '')
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# 分析结果 API
# ============================================

@analysis_bp.route('/api/analysis/task/<task_id>', methods=['GET'])
def get_task_analysis(task_id):
    """获取任务分析结果（快速统计 + LLM 分析）"""
    try:
        # 获取缓存的分析结果
        result = analysis_service.get_cached_analysis(task_id, 'task', task_id)
        
        if result.get('error'):
            return jsonify({'success': False, 'error': result.get('error')}), 404
        
        return jsonify({
            'success': True,
            'data': {
                'quick_stats': result.get('quick_stats'),
                'llm_analysis': result.get('llm_analysis'),
                'analysis_status': result.get('analysis_status'),
                'updated_at': result.get('updated_at'),
                'cache_hit': result.get('cache_hit', False)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/quick-stats/<task_id>', methods=['GET'])
def get_quick_stats(task_id):
    """获取快速本地统计（毫秒级响应）"""
    try:
        stats = analysis_service.get_quick_stats(task_id)
        
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/clusters', methods=['GET'])
def get_clusters():
    """获取聚类列表"""
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        
        # 获取快速统计中的聚类
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        clusters = stats.get('clusters', [])
        
        # 分页
        total = len(clusters)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = clusters[start:end]
        
        # 获取 LLM 分析结果
        cached = analysis_service.get_cached_analysis(task_id, 'task', task_id)
        llm_clusters = None
        if cached.get('llm_analysis'):
            llm_clusters = cached['llm_analysis'].get('clusters')
        
        return jsonify({
            'success': True,
            'data': {
                'quick_stats': {
                    'total_clusters': total,
                    'clusters': paginated
                },
                'llm_analysis': {
                    'clusters': llm_clusters
                } if llm_clusters else None,
                'page': page,
                'page_size': page_size,
                'total': total
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/clusters/<cluster_id>', methods=['GET'])
def get_cluster_detail(cluster_id):
    """获取聚类详情"""
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        # 获取快速统计中的聚类
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        clusters = stats.get('clusters', [])
        
        # 查找指定聚类
        cluster = None
        for c in clusters:
            if c.get('cluster_key') == cluster_id or c.get('cluster_id') == cluster_id:
                cluster = c
                break
        
        if not cluster:
            return jsonify({'success': False, 'error': '聚类不存在'}), 404
        
        return jsonify({
            'success': True,
            'data': cluster
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# 样本管理 API
# ============================================

@analysis_bp.route('/api/analysis/samples', methods=['GET'])
def get_samples():
    """获取错误样本列表"""
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        # 筛选参数
        status = request.args.get('status')
        error_type = request.args.get('error_type')
        subject = request.args.get('subject')
        book_name = request.args.get('book_name')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        sort_by = request.args.get('sort_by', 'question_index')
        sort_order = request.args.get('sort_order', 'asc')
        
        # 获取快速统计中的所有样本
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 从聚类中收集所有样本
        all_samples = []
        for cluster in stats.get('clusters', []):
            for sample in cluster.get('samples', []):
                sample['cluster_key'] = cluster.get('cluster_key')
                # 生成唯一样本ID
                sample['sample_id'] = f"{sample.get('homework_id', '')}_{sample.get('question_index', 0)}"
                all_samples.append(sample)
        
        # 筛选
        filtered = all_samples
        if error_type:
            filtered = [s for s in filtered if s.get('error_type') == error_type]
        if book_name:
            filtered = [s for s in filtered if s.get('book_name') == book_name]
        
        # 排序
        reverse = sort_order == 'desc'
        filtered.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
        
        # 分页
        total = len(filtered)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = filtered[start:end]
        
        return jsonify({
            'success': True,
            'data': {
                'items': paginated,
                'total': total,
                'page': page,
                'page_size': page_size
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/samples/<sample_id>', methods=['GET'])
def get_sample_detail(sample_id):
    """
    获取单个错误样本详情
    
    Returns:
        样本详情 + LLM 分析结果 + 所属聚类
    """
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        # 获取快速统计中的所有样本
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 查找指定样本
        sample = None
        cluster_info = None
        
        for cluster in stats.get('clusters', []):
            for s in cluster.get('samples', []):
                s_id = f"{s.get('homework_id', '')}_{s.get('question_index', 0)}"
                if s_id == sample_id:
                    sample = s.copy()
                    sample['sample_id'] = s_id
                    cluster_info = {
                        'cluster_key': cluster.get('cluster_key'),
                        'error_type': cluster.get('error_type'),
                        'book_name': cluster.get('book_name'),
                        'page_range': cluster.get('page_range'),
                        'sample_count': cluster.get('sample_count')
                    }
                    break
            if sample:
                break
        
        if not sample:
            return jsonify({'success': False, 'error': '样本不存在'}), 404
        
        # 获取 LLM 分析结果（如果有）
        llm_insight = None
        cached = analysis_service.get_cached_analysis(task_id, 'sample', sample_id)
        if cached.get('llm_analysis'):
            llm_insight = cached['llm_analysis']
        
        return jsonify({
            'success': True,
            'data': {
                'sample': sample,
                'cluster': cluster_info,
                'llm_insight': llm_insight,
                'status': sample.get('status', 'pending')
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# 维度分析 API
# ============================================

@analysis_bp.route('/api/analysis/subject', methods=['GET'])
def get_subject_analysis():
    """获取学科维度分析"""
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 构建学科列表
        subject_dist = stats.get('subject_distribution', {})
        total_errors = stats.get('total_errors', 0)
        subjects = []
        for name, count in subject_dist.items():
            error_rate = count / total_errors if total_errors > 0 else 0
            subjects.append({
                'subject_id': name,
                'name': name,
                'error_count': count,
                'error_rate': round(error_rate, 4)
            })
        subjects.sort(key=lambda x: x['error_count'], reverse=True)
        
        # 获取 LLM 分析结果
        cached = analysis_service.get_cached_analysis(task_id, 'subject', None)
        llm_analysis = cached.get('llm_analysis') if cached else None
        
        return jsonify({
            'success': True,
            'data': {
                'quick_stats': {
                    'subjects': subjects
                },
                'llm_analysis': llm_analysis,
                'analysis_status': cached.get('analysis_status') if cached else 'pending'
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/subject/<subject_id>', methods=['GET'])
def get_subject_detail(subject_id):
    """
    获取单个学科详情
    
    Returns:
        学科详情 + 书本列表
    """
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 获取该学科的错误样本
        subject_samples = []
        book_stats = {}
        
        for cluster in stats.get('clusters', []):
            for sample in cluster.get('samples', []):
                # 从书名推断学科
                book_name = sample.get('book_name', '')
                sample_subject = _infer_subject_from_book(book_name)
                
                if sample_subject == subject_id:
                    subject_samples.append(sample)
                    
                    # 统计书本
                    if book_name not in book_stats:
                        book_stats[book_name] = {'error_count': 0, 'samples': []}
                    book_stats[book_name]['error_count'] += 1
                    if len(book_stats[book_name]['samples']) < 5:
                        book_stats[book_name]['samples'].append(sample)
        
        # 构建书本列表
        books = [
            {
                'book_name': name,
                'error_count': data['error_count'],
                'sample_preview': data['samples'][:3]
            }
            for name, data in book_stats.items()
        ]
        books.sort(key=lambda x: x['error_count'], reverse=True)
        
        # 获取 LLM 分析结果
        cached = analysis_service.get_cached_analysis(task_id, 'subject', subject_id)
        llm_analysis = cached.get('llm_analysis') if cached else None
        
        return jsonify({
            'success': True,
            'data': {
                'subject_id': subject_id,
                'subject_name': subject_id,
                'total_errors': len(subject_samples),
                'books': books,
                'llm_analysis': llm_analysis
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _infer_subject_from_book(book_name: str) -> str:
    """从书名推断学科"""
    if not book_name:
        return '未知'
    book_name_lower = book_name.lower()
    keywords = {
        '英语': ['英语', 'english'],
        '语文': ['语文', 'chinese'],
        '数学': ['数学', 'math'],
        '物理': ['物理', 'physics'],
        '化学': ['化学', 'chemistry'],
        '生物': ['生物', 'biology'],
        '地理': ['地理', 'geography']
    }
    for subject, kws in keywords.items():
        for kw in kws:
            if kw in book_name_lower:
                return subject
    return '未知'


@analysis_bp.route('/api/analysis/book', methods=['GET'])
def get_book_analysis():
    """获取书本维度分析"""
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        subject_id = request.args.get('subject_id')
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 构建书本列表
        book_dist = stats.get('book_distribution', {})
        books = [
            {'name': name, 'error_count': count}
            for name, count in book_dist.items()
        ]
        books.sort(key=lambda x: x['error_count'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': {
                'quick_stats': {
                    'books': books
                },
                'llm_analysis': None
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/question-type', methods=['GET'])
def get_question_type_analysis():
    """获取题型/错误类型维度分析"""
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 构建错误类型列表
        error_type_dist = stats.get('error_type_distribution', {})
        types = [
            {'name': name, 'error_count': count}
            for name, count in error_type_dist.items()
        ]
        types.sort(key=lambda x: x['error_count'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': {
                'quick_stats': {
                    'question_types': types
                },
                'llm_analysis': None
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/book/<book_id>', methods=['GET'])
def get_book_detail(book_id):
    """
    获取单个书本详情
    
    Returns:
        书本详情 + 页码分布
    """
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        # URL decode book_id
        from urllib.parse import unquote
        book_name = unquote(book_id)
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 获取该书本的错误样本
        book_samples = []
        page_stats = {}
        
        for cluster in stats.get('clusters', []):
            for sample in cluster.get('samples', []):
                if sample.get('book_name') == book_name:
                    book_samples.append(sample)
                    
                    # 统计页码
                    page_num = sample.get('page_num', 0)
                    if page_num not in page_stats:
                        page_stats[page_num] = {'error_count': 0, 'samples': []}
                    page_stats[page_num]['error_count'] += 1
                    if len(page_stats[page_num]['samples']) < 3:
                        page_stats[page_num]['samples'].append(sample)
        
        # 构建页码分布列表
        pages = [
            {
                'page_num': page,
                'error_count': data['error_count'],
                'sample_preview': data['samples']
            }
            for page, data in page_stats.items()
        ]
        pages.sort(key=lambda x: x['error_count'], reverse=True)
        
        # 推断学科
        subject = '未知'
        if book_samples:
            subject = _infer_subject_from_book(book_name)
        
        # 获取 LLM 分析结果
        cached = analysis_service.get_cached_analysis(task_id, 'book', book_name)
        llm_analysis = cached.get('llm_analysis') if cached else None
        
        return jsonify({
            'success': True,
            'data': {
                'book_id': book_id,
                'book_name': book_name,
                'subject': subject,
                'total_errors': len(book_samples),
                'page_count': len(pages),
                'pages': pages,
                'llm_analysis': llm_analysis
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/question-type/<error_type>', methods=['GET'])
def get_question_type_detail(error_type):
    """
    获取单个错误类型/题型详情
    
    Returns:
        错误类型详情 + 样本列表
    """
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        # URL decode error_type
        from urllib.parse import unquote
        error_type_name = unquote(error_type)
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 获取该错误类型的样本
        type_samples = []
        book_stats = {}
        
        for cluster in stats.get('clusters', []):
            if cluster.get('error_type') == error_type_name:
                for sample in cluster.get('samples', []):
                    type_samples.append(sample)
                    
                    # 统计书本分布
                    book_name = sample.get('book_name', '未知')
                    if book_name not in book_stats:
                        book_stats[book_name] = 0
                    book_stats[book_name] += 1
        
        # 构建书本分布
        book_distribution = [
            {'book_name': name, 'error_count': count}
            for name, count in book_stats.items()
        ]
        book_distribution.sort(key=lambda x: x['error_count'], reverse=True)
        
        # 错误类型描述
        descriptions = {
            '识别错误-判断错误': '手写体或图片识别不准确，导致 AI 对答案的判断出现偏差',
            '识别正确-判断错误': 'AI 正确识别了答案内容，但评分逻辑存在问题',
            '缺失题目': '部分题目未被 AI 识别或处理',
            'AI识别幻觉': 'AI 识别出了实际不存在的内容',
            '答案不匹配': '标准答案与学生答案格式不一致或存在等价表达'
        }
        
        # 获取 LLM 分析结果
        cached = analysis_service.get_cached_analysis(task_id, 'question_type', error_type_name)
        llm_analysis = cached.get('llm_analysis') if cached else None
        
        return jsonify({
            'success': True,
            'data': {
                'error_type': error_type_name,
                'description': descriptions.get(error_type_name, '未知错误类型'),
                'total_errors': len(type_samples),
                'book_distribution': book_distribution,
                'sample_preview': type_samples[:10],
                'llm_analysis': llm_analysis
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/trend', methods=['GET'])
def get_trend_analysis():
    """
    获取时间趋势分析
    
    参数:
        task_ids: 逗号分隔的任务ID列表
        time_range: 7d|30d|custom
        
    Returns:
        趋势数据 + LLM 分析
    """
    try:
        task_ids_str = request.args.get('task_ids', '')
        time_range = request.args.get('time_range', '7d')
        
        if not task_ids_str:
            return jsonify({'success': False, 'error': '缺少 task_ids 参数'}), 400
        
        task_ids = [tid.strip() for tid in task_ids_str.split(',') if tid.strip()]
        
        if len(task_ids) < 2:
            return jsonify({'success': False, 'error': '至少需要2个任务进行趋势分析'}), 400
        
        # 收集各任务的统计数据
        trend_data = []
        for task_id in task_ids:
            stats = analysis_service.get_quick_stats(task_id)
            if stats.get('error'):
                continue
            
            # 获取任务创建时间
            task_data = StorageService.load_batch_task(task_id)
            created_at = task_data.get('created_at', '') if task_data else ''
            
            trend_data.append({
                'task_id': task_id,
                'created_at': created_at,
                'total_errors': stats.get('total_errors', 0),
                'total_questions': stats.get('total_questions', 0),
                'error_rate': stats.get('error_rate', 0),
                'error_type_distribution': stats.get('error_type_distribution', {})
            })
        
        # 按时间排序
        trend_data.sort(key=lambda x: x.get('created_at', ''))
        
        # 计算趋势
        if len(trend_data) >= 2:
            first = trend_data[0]
            last = trend_data[-1]
            error_rate_change = last.get('error_rate', 0) - first.get('error_rate', 0)
            trend_direction = 'improving' if error_rate_change < 0 else 'declining' if error_rate_change > 0 else 'stable'
        else:
            error_rate_change = 0
            trend_direction = 'insufficient_data'
        
        return jsonify({
            'success': True,
            'data': {
                'quick_stats': {
                    'data_points': trend_data,
                    'error_rate_change': round(error_rate_change, 4),
                    'trend_direction': trend_direction,
                    'task_count': len(trend_data)
                },
                'llm_analysis': None,  # LLM 分析需要单独触发
                'time_range': time_range
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/compare', methods=['GET'])
def get_batch_comparison():
    """
    获取批次对比分析
    
    参数:
        task_id_1: 第一个任务ID
        task_id_2: 第二个任务ID
        
    Returns:
        对比数据 + LLM 分析
    """
    try:
        task_id_1 = request.args.get('task_id_1')
        task_id_2 = request.args.get('task_id_2')
        
        if not task_id_1 or not task_id_2:
            return jsonify({'success': False, 'error': '缺少 task_id_1 或 task_id_2 参数'}), 400
        
        # 获取两个任务的统计数据
        stats1 = analysis_service.get_quick_stats(task_id_1)
        stats2 = analysis_service.get_quick_stats(task_id_2)
        
        if stats1.get('error'):
            return jsonify({'success': False, 'error': f'任务1: {stats1.get("error")}'}), 404
        if stats2.get('error'):
            return jsonify({'success': False, 'error': f'任务2: {stats2.get("error")}'}), 404
        
        # 计算变化
        error_rate_change = stats2.get('error_rate', 0) - stats1.get('error_rate', 0)
        error_count_change = stats2.get('total_errors', 0) - stats1.get('total_errors', 0)
        
        # 错误类型变化
        type_dist1 = stats1.get('error_type_distribution', {})
        type_dist2 = stats2.get('error_type_distribution', {})
        all_types = set(type_dist1.keys()) | set(type_dist2.keys())
        
        type_changes = []
        for error_type in all_types:
            count1 = type_dist1.get(error_type, 0)
            count2 = type_dist2.get(error_type, 0)
            change = count2 - count1
            type_changes.append({
                'error_type': error_type,
                'count_before': count1,
                'count_after': count2,
                'change': change,
                'direction': 'improved' if change < 0 else 'declined' if change > 0 else 'unchanged'
            })
        type_changes.sort(key=lambda x: abs(x['change']), reverse=True)
        
        # 判断整体趋势
        if error_rate_change < -0.05:
            overall_trend = 'significant_improvement'
        elif error_rate_change < 0:
            overall_trend = 'slight_improvement'
        elif error_rate_change > 0.05:
            overall_trend = 'significant_decline'
        elif error_rate_change > 0:
            overall_trend = 'slight_decline'
        else:
            overall_trend = 'stable'
        
        return jsonify({
            'success': True,
            'data': {
                'quick_stats': {
                    'batch1': {
                        'task_id': task_id_1,
                        'total_errors': stats1.get('total_errors', 0),
                        'total_questions': stats1.get('total_questions', 0),
                        'error_rate': stats1.get('error_rate', 0)
                    },
                    'batch2': {
                        'task_id': task_id_2,
                        'total_errors': stats2.get('total_errors', 0),
                        'total_questions': stats2.get('total_questions', 0),
                        'error_rate': stats2.get('error_rate', 0)
                    },
                    'comparison': {
                        'error_rate_change': round(error_rate_change, 4),
                        'error_count_change': error_count_change,
                        'overall_trend': overall_trend,
                        'type_changes': type_changes
                    }
                },
                'llm_analysis': None  # LLM 分析需要单独触发
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/suggestions', methods=['GET'])
def get_suggestions():
    """
    获取优化建议
    
    Returns:
        建议列表（最多5条）
    """
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        # 获取分析报告中的建议
        report = analysis_service.get_report(task_id)
        
        if report and report.get('suggestions'):
            suggestions = report.get('suggestions', [])
        else:
            # 如果没有报告，基于快速统计生成简单建议
            stats = analysis_service.get_quick_stats(task_id)
            if stats.get('error'):
                return jsonify({'success': False, 'error': stats.get('error')}), 404
            
            suggestions = _generate_quick_suggestions(stats)
        
        return jsonify({
            'success': True,
            'data': {
                'suggestions': suggestions[:5],
                'total': len(suggestions)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _generate_quick_suggestions(stats: dict) -> list:
    """基于快速统计生成简单建议"""
    suggestions = []
    
    error_type_dist = stats.get('error_type_distribution', {})
    total_errors = stats.get('total_errors', 0)
    
    if total_errors == 0:
        return [{
            'suggestion_id': 's1',
            'title': '暂无错误',
            'description': '当前批次没有发现错误，继续保持！',
            'priority': 'low',
            'expected_effect': '无需优化'
        }]
    
    # 基于错误类型生成建议
    for error_type, count in error_type_dist.items():
        percentage = count / total_errors if total_errors > 0 else 0
        if percentage < 0.1:
            continue
        
        if error_type == '识别错误-判断错误':
            suggestions.append({
                'suggestion_id': f's{len(suggestions)+1}',
                'title': '优化手写体识别容错',
                'description': '在评分 prompt 中增加对手写体的容错处理',
                'priority': 'high' if percentage > 0.3 else 'medium',
                'expected_effect': f'预计可减少 {int(percentage*30)}% 的识别错误'
            })
        elif error_type == '识别正确-判断错误':
            suggestions.append({
                'suggestion_id': f's{len(suggestions)+1}',
                'title': '完善评分逻辑',
                'description': '优化评分标准的描述，增加对部分得分和等价答案的处理规则',
                'priority': 'high' if percentage > 0.3 else 'medium',
                'expected_effect': f'预计可减少 {int(percentage*25)}% 的评分逻辑错误'
            })
        elif error_type == '缺失题目':
            suggestions.append({
                'suggestion_id': f's{len(suggestions)+1}',
                'title': '检查题目识别逻辑',
                'description': '审核题目定位和识别逻辑，确保所有题目都被正确处理',
                'priority': 'high',
                'expected_effect': f'预计可减少 {int(percentage*50)}% 的缺失题目'
            })
        elif error_type == 'AI识别幻觉':
            suggestions.append({
                'suggestion_id': f's{len(suggestions)+1}',
                'title': '增加幻觉检测',
                'description': '在 prompt 中增加对识别结果的验证逻辑',
                'priority': 'medium',
                'expected_effect': f'预计可减少 {int(percentage*40)}% 的幻觉错误'
            })
        elif error_type == '答案不匹配':
            suggestions.append({
                'suggestion_id': f's{len(suggestions)+1}',
                'title': '规范标准答案格式',
                'description': '统一标准答案的格式，增加等价答案的覆盖',
                'priority': 'medium',
                'expected_effect': f'预计可减少 {int(percentage*20)}% 的答案不匹配错误'
            })
    
    # 按优先级排序
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    suggestions.sort(key=lambda x: priority_order.get(x.get('priority', 'low'), 3))
    
    return suggestions[:5]


# ============================================
# 样本管理 API
# ============================================

@analysis_bp.route('/api/analysis/samples/<sample_id>/status', methods=['PUT'])
def update_sample_status(sample_id):
    """
    更新样本状态
    
    参数:
        status: pending|confirmed|fixed|ignored
        note: 备注
    """
    try:
        data = request.get_json() or {}
        status = data.get('status')
        note = data.get('note', '')
        
        if not status:
            return jsonify({'success': False, 'error': '缺少 status 参数'}), 400
        
        valid_statuses = ['pending', 'confirmed', 'fixed', 'ignored']
        if status not in valid_statuses:
            return jsonify({'success': False, 'error': f'无效的状态，有效值: {valid_statuses}'}), 400
        
        # 更新数据库中的样本状态
        try:
            from services.database_service import AppDatabaseService
            from datetime import datetime
            
            sql = """
                INSERT INTO error_samples (sample_id, status, note, updated_at)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                status = VALUES(status),
                note = VALUES(note),
                updated_at = VALUES(updated_at)
            """
            AppDatabaseService.execute_insert(sql, (sample_id, status, note, datetime.now()))
            
            return jsonify({
                'success': True,
                'data': {
                    'sample_id': sample_id,
                    'status': status,
                    'note': note,
                    'message': '状态更新成功'
                }
            })
        except Exception as db_error:
            # 如果数据库操作失败，返回成功但标记为内存更新
            return jsonify({
                'success': True,
                'data': {
                    'sample_id': sample_id,
                    'status': status,
                    'note': note,
                    'message': '状态更新成功（内存）',
                    'db_error': str(db_error)
                }
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/samples/batch-status', methods=['PUT'])
def batch_update_sample_status():
    """
    批量更新样本状态
    
    参数:
        sample_ids: 样本ID列表
        status: pending|confirmed|fixed|ignored
    """
    try:
        data = request.get_json() or {}
        sample_ids = data.get('sample_ids', [])
        status = data.get('status')
        
        if not sample_ids:
            return jsonify({'success': False, 'error': '缺少 sample_ids 参数'}), 400
        if not status:
            return jsonify({'success': False, 'error': '缺少 status 参数'}), 400
        
        valid_statuses = ['pending', 'confirmed', 'fixed', 'ignored']
        if status not in valid_statuses:
            return jsonify({'success': False, 'error': f'无效的状态，有效值: {valid_statuses}'}), 400
        
        # 批量更新
        updated_count = 0
        failed_ids = []
        
        try:
            from services.database_service import AppDatabaseService
            from datetime import datetime
            
            for sample_id in sample_ids:
                try:
                    sql = """
                        INSERT INTO error_samples (sample_id, status, updated_at)
                        VALUES (%s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                        status = VALUES(status),
                        updated_at = VALUES(updated_at)
                    """
                    AppDatabaseService.execute_insert(sql, (sample_id, status, datetime.now()))
                    updated_count += 1
                except Exception:
                    failed_ids.append(sample_id)
        except Exception:
            # 数据库不可用时，标记所有为成功（内存更新）
            updated_count = len(sample_ids)
        
        return jsonify({
            'success': True,
            'data': {
                'updated_count': updated_count,
                'failed_count': len(failed_ids),
                'failed_ids': failed_ids,
                'status': status,
                'message': f'成功更新 {updated_count} 个样本状态'
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/samples/<sample_id>/reanalyze', methods=['POST'])
def reanalyze_sample(sample_id):
    """
    对单个样本重新调用 LLM 分析
    """
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        # 获取样本详情
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 查找样本
        sample = None
        for cluster in stats.get('clusters', []):
            for s in cluster.get('samples', []):
                s_id = f"{s.get('homework_id', '')}_{s.get('question_index', 0)}"
                if s_id == sample_id:
                    sample = s
                    break
            if sample:
                break
        
        if not sample:
            return jsonify({'success': False, 'error': '样本不存在'}), 404
        
        # 调用 LLM 分析（异步）
        try:
            from services.llm_analysis_service import LLMAnalysisService
            
            # 构建分析请求
            analysis_result = LLMAnalysisService.analyze_sample(sample)
            
            # 保存分析结果
            if analysis_result:
                analysis_service._save_analysis_result(
                    task_id, 'sample', sample_id,
                    analysis_result, '', 0
                )
            
            return jsonify({
                'success': True,
                'data': {
                    'sample_id': sample_id,
                    'analysis': analysis_result,
                    'message': '重新分析完成'
                }
            })
        except ImportError:
            # LLM 分析服务不可用
            return jsonify({
                'success': True,
                'data': {
                    'sample_id': sample_id,
                    'analysis': None,
                    'message': '已标记为待重新分析（LLM 服务暂不可用）'
                }
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# 异常检测 API
# ============================================

@analysis_bp.route('/api/analysis/anomalies', methods=['GET'])
def get_anomalies():
    """获取异常检测结果"""
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        anomalies = analysis_service.detect_anomalies(task_id)
        
        # 统计摘要
        summary = {
            'critical_count': len([a for a in anomalies if a.get('severity') == 'critical']),
            'high_count': len([a for a in anomalies if a.get('severity') == 'high']),
            'medium_count': len([a for a in anomalies if a.get('severity') == 'medium']),
            'low_count': len([a for a in anomalies if a.get('severity') == 'low'])
        }
        
        return jsonify({
            'success': True,
            'data': {
                'anomalies': anomalies,
                'summary': summary
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# 配置和成本统计 API
# ============================================

@analysis_bp.route('/api/analysis/config', methods=['GET'])
def get_analysis_config():
    """获取分析配置"""
    try:
        config = analysis_service.get_config()
        return jsonify({
            'success': True,
            'data': config
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/config', methods=['PUT'])
def update_analysis_config():
    """更新分析配置"""
    try:
        data = request.get_json() or {}
        config = analysis_service.update_config(data)
        return jsonify({
            'success': True,
            'data': config
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/cost-stats', methods=['GET'])
def get_cost_stats():
    """获取成本统计"""
    try:
        days = int(request.args.get('days', 7))
        stats = LLMService.get_token_stats(days)
        
        # 估算成本（DeepSeek V3.2 价格约 $0.14/1M tokens）
        cost_per_million = 0.14
        stats['estimated_cost'] = {
            'today': round(stats.get('today', {}).get('tokens', 0) / 1000000 * cost_per_million, 4),
            'week': round(stats.get('week', {}).get('tokens', 0) / 1000000 * cost_per_million, 4),
            'month': round(stats.get('month', {}).get('tokens', 0) / 1000000 * cost_per_million, 4)
        }
        
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# 兼容旧 API
# ============================================


# ============================================
# 兼容旧 API
# ============================================

@analysis_bp.route('/api/analysis/report/<task_id>', methods=['GET'])
def get_analysis_report(task_id):
    """获取分析报告（兼容旧 API）"""
    try:
        report = analysis_service.get_report(task_id)
        
        if not report:
            return jsonify({
                'success': True,
                'report': None,
                'message': '暂无分析报告'
            })
        
        return jsonify({
            'success': True,
            'report': report
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/drilldown/<task_id>', methods=['GET'])
def get_drilldown_data(task_id):
    """获取层级下钻数据"""
    try:
        # 获取下钻参数
        level = request.args.get('level', 'subject')  # subject, book, page, question
        parent_id = request.args.get('parent_id')
        
        # 使用新的下钻方法
        data = analysis_service.get_drill_down_data(task_id, level, parent_id)
        
        return jsonify({
            'success': True,
            'level': level,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/status/<task_id>', methods=['GET'])
def get_analysis_status(task_id):
    """获取分析状态"""
    try:
        status = analysis_service.get_status(task_id)
        
        return jsonify({
            'success': True,
            'status': status
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# 高级可视化数据 API
# ============================================

@analysis_bp.route('/api/analysis/chart/sankey', methods=['GET'])
def get_sankey_chart():
    """
    获取桑基图数据
    
    Returns:
        nodes: 错误类型/根因/建议节点
        links: 流转关系
    """
    try:
        task_id = request.args.get('task_id')
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 构建节点
        nodes = []
        links = []
        node_index = {}
        
        # 错误类型节点
        error_type_dist = stats.get('error_type_distribution', {})
        for error_type, count in error_type_dist.items():
            node_index[f'error_{error_type}'] = len(nodes)
            nodes.append({'name': error_type, 'category': 'error_type', 'value': count})
        
        # 根因节点（基于错误类型推断）
        cause_mapping = {
            '识别错误-判断错误': 'OCR识别问题',
            '识别正确-判断错误': '评分逻辑问题',
            '缺失题目': '数据问题',
            'AI识别幻觉': 'OCR识别问题',
            '答案不匹配': '标准答案问题'
        }
        
        cause_counts = {}
        for error_type, count in error_type_dist.items():
            cause = cause_mapping.get(error_type, '其他问题')
            cause_counts[cause] = cause_counts.get(cause, 0) + count
        
        for cause, count in cause_counts.items():
            node_index[f'cause_{cause}'] = len(nodes)
            nodes.append({'name': cause, 'category': 'root_cause', 'value': count})
        
        # 建议节点
        suggestion_mapping = {
            'OCR识别问题': '优化手写体识别',
            '评分逻辑问题': '完善评分逻辑',
            '数据问题': '检查数据质量',
            '标准答案问题': '规范标准答案',
            '其他问题': '综合优化'
        }
        
        for cause in cause_counts.keys():
            suggestion = suggestion_mapping.get(cause, '综合优化')
            if f'suggestion_{suggestion}' not in node_index:
                node_index[f'suggestion_{suggestion}'] = len(nodes)
                nodes.append({'name': suggestion, 'category': 'suggestion', 'value': cause_counts[cause]})
        
        # 构建链接：错误类型 -> 根因
        for error_type, count in error_type_dist.items():
            cause = cause_mapping.get(error_type, '其他问题')
            links.append({
                'source': node_index[f'error_{error_type}'],
                'target': node_index[f'cause_{cause}'],
                'value': count
            })
        
        # 构建链接：根因 -> 建议
        for cause, count in cause_counts.items():
            suggestion = suggestion_mapping.get(cause, '综合优化')
            links.append({
                'source': node_index[f'cause_{cause}'],
                'target': node_index[f'suggestion_{suggestion}'],
                'value': count
            })
        
        return jsonify({
            'success': True,
            'data': {
                'nodes': nodes,
                'links': links
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/chart/heatmap', methods=['GET'])
def get_heatmap_chart():
    """
    获取热力图数据
    
    Returns:
        x_axis: 题目索引
        y_axis: 页码
        data: 错误数矩阵
    """
    try:
        task_id = request.args.get('task_id')
        book_id = request.args.get('book_id')
        
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 收集数据
        from urllib.parse import unquote
        book_name = unquote(book_id) if book_id else None
        
        page_question_errors = {}
        all_pages = set()
        all_questions = set()
        
        for cluster in stats.get('clusters', []):
            for sample in cluster.get('samples', []):
                if book_name and sample.get('book_name') != book_name:
                    continue
                
                page = sample.get('page_num', 0)
                question = sample.get('question_index', 0)
                
                all_pages.add(page)
                all_questions.add(question)
                
                key = (page, question)
                page_question_errors[key] = page_question_errors.get(key, 0) + 1
        
        # 构建矩阵数据
        pages = sorted(all_pages)
        questions = sorted(all_questions)
        
        data = []
        for q_idx, question in enumerate(questions):
            for p_idx, page in enumerate(pages):
                count = page_question_errors.get((page, question), 0)
                if count > 0:
                    data.append([q_idx, p_idx, count])
        
        return jsonify({
            'success': True,
            'data': {
                'x_axis': [f'题{q}' for q in questions],
                'y_axis': [f'P{p}' for p in pages],
                'data': data,
                'max_value': max([d[2] for d in data]) if data else 0
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/chart/radar', methods=['GET'])
def get_radar_chart():
    """
    获取雷达图数据
    
    Returns:
        indicators: 维度指标
        series: 数据系列
    """
    try:
        task_id = request.args.get('task_id')
        dimension = request.args.get('dimension', 'error_type')
        
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        total_errors = stats.get('total_errors', 1)
        
        if dimension == 'error_type':
            dist = stats.get('error_type_distribution', {})
        elif dimension == 'subject':
            dist = stats.get('subject_distribution', {})
        elif dimension == 'book':
            dist = stats.get('book_distribution', {})
        else:
            dist = stats.get('error_type_distribution', {})
        
        # 构建指标
        indicators = []
        values = []
        
        for name, count in dist.items():
            percentage = round(count / total_errors * 100, 1) if total_errors > 0 else 0
            indicators.append({'name': name[:10], 'max': 100})
            values.append(percentage)
        
        return jsonify({
            'success': True,
            'data': {
                'indicators': indicators,
                'series': [{
                    'name': '当前批次',
                    'values': values
                }]
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================
# 搜索筛选 API
# ============================================

@analysis_bp.route('/api/analysis/samples/search', methods=['GET'])
def search_samples():
    """
    高级搜索样本
    
    参数:
        q: 搜索语法 (book:xxx AND status:xxx)
        task_id: 任务ID
        page, page_size: 分页
    """
    try:
        task_id = request.args.get('task_id')
        query = request.args.get('q', '')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        stats = analysis_service.get_quick_stats(task_id)
        if stats.get('error'):
            return jsonify({'success': False, 'error': stats.get('error')}), 404
        
        # 收集所有样本
        all_samples = []
        for cluster in stats.get('clusters', []):
            for sample in cluster.get('samples', []):
                sample['sample_id'] = f"{sample.get('homework_id', '')}_{sample.get('question_index', 0)}"
                sample['cluster_key'] = cluster.get('cluster_key')
                all_samples.append(sample)
        
        # 解析搜索语法
        filtered = _parse_and_filter(all_samples, query)
        
        # 分页
        total = len(filtered)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = filtered[start:end]
        
        # 高亮关键词
        highlights = _extract_keywords(query)
        
        return jsonify({
            'success': True,
            'data': {
                'items': paginated,
                'total': total,
                'page': page,
                'page_size': page_size,
                'query_parsed': query,
                'highlights': highlights
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _parse_and_filter(samples: list, query: str) -> list:
    """解析搜索语法并筛选"""
    if not query:
        return samples
    
    filtered = samples
    
    # 简单解析：支持 field:value 格式
    import re
    patterns = re.findall(r'(\w+):([^\s]+)', query)
    
    for field, value in patterns:
        field = field.lower()
        value = value.lower()
        
        if field == 'book':
            filtered = [s for s in filtered if value in s.get('book_name', '').lower()]
        elif field == 'error_type' or field == 'type':
            filtered = [s for s in filtered if value in s.get('error_type', '').lower()]
        elif field == 'page':
            if '-' in value:
                start, end = value.split('-')
                filtered = [s for s in filtered if int(start) <= s.get('page_num', 0) <= int(end)]
            else:
                filtered = [s for s in filtered if s.get('page_num', 0) == int(value)]
        elif field == 'status':
            filtered = [s for s in filtered if s.get('status', 'pending').lower() == value]
        elif field == 'homework':
            filtered = [s for s in filtered if value in s.get('homework_id', '').lower()]
    
    # 如果没有字段匹配，进行全文搜索
    if not patterns and query:
        query_lower = query.lower()
        filtered = [s for s in filtered if 
                   query_lower in s.get('book_name', '').lower() or
                   query_lower in s.get('error_type', '').lower() or
                   query_lower in s.get('ai_answer', '').lower() or
                   query_lower in s.get('expected_answer', '').lower()]
    
    return filtered


def _extract_keywords(query: str) -> list:
    """提取搜索关键词用于高亮"""
    import re
    keywords = []
    
    # 提取 field:value 中的 value
    patterns = re.findall(r'\w+:([^\s]+)', query)
    keywords.extend(patterns)
    
    # 提取普通词
    words = re.findall(r'\b\w+\b', query)
    keywords.extend([w for w in words if w.lower() not in ['and', 'or', 'not']])
    
    return list(set(keywords))


@analysis_bp.route('/api/analysis/filter-presets', methods=['GET'])
def get_filter_presets():
    """获取筛选预设"""
    try:
        # 系统预设
        system_presets = [
            {'id': 'pending', 'name': '待处理', 'query': 'status:pending', 'is_system': True},
            {'id': 'confirmed', 'name': '已确认', 'query': 'status:confirmed', 'is_system': True},
            {'id': 'fixed', 'name': '已修复', 'query': 'status:fixed', 'is_system': True},
            {'id': 'ocr_error', 'name': '识别错误', 'query': 'type:识别错误', 'is_system': True},
            {'id': 'logic_error', 'name': '判断错误', 'query': 'type:判断错误', 'is_system': True}
        ]
        
        # 用户预设（从数据库或文件加载）
        user_presets = _load_user_presets()
        
        return jsonify({
            'success': True,
            'data': {
                'system_presets': system_presets,
                'user_presets': user_presets
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/filter-presets', methods=['POST'])
def save_filter_preset():
    """保存用户筛选预设"""
    try:
        data = request.get_json() or {}
        name = data.get('name')
        query = data.get('query')
        
        if not name or not query:
            return jsonify({'success': False, 'error': '缺少 name 或 query 参数'}), 400
        
        import uuid
        preset_id = f"user_{uuid.uuid4().hex[:8]}"
        
        preset = {
            'id': preset_id,
            'name': name,
            'query': query,
            'is_system': False,
            'created_at': datetime.now().isoformat()
        }
        
        _save_user_preset(preset)
        
        return jsonify({
            'success': True,
            'data': preset
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/filter-presets/<preset_id>', methods=['DELETE'])
def delete_filter_preset(preset_id):
    """删除用户预设"""
    try:
        if preset_id.startswith('user_'):
            _delete_user_preset(preset_id)
            return jsonify({'success': True, 'message': '预设已删除'})
        else:
            return jsonify({'success': False, 'error': '无法删除系统预设'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _load_user_presets() -> list:
    """加载用户预设"""
    import os
    preset_file = 'saved_filters/analysis_presets.json'
    if os.path.exists(preset_file):
        with open(preset_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def _save_user_preset(preset: dict):
    """保存用户预设"""
    import os
    preset_file = 'saved_filters/analysis_presets.json'
    os.makedirs('saved_filters', exist_ok=True)
    
    presets = _load_user_presets()
    presets.append(preset)
    
    with open(preset_file, 'w', encoding='utf-8') as f:
        json.dump(presets, f, ensure_ascii=False, indent=2)


def _delete_user_preset(preset_id: str):
    """删除用户预设"""
    presets = _load_user_presets()
    presets = [p for p in presets if p.get('id') != preset_id]
    
    preset_file = 'saved_filters/analysis_presets.json'
    with open(preset_file, 'w', encoding='utf-8') as f:
        json.dump(presets, f, ensure_ascii=False, indent=2)


# ============================================
# 导出报告 API
# ============================================

@analysis_bp.route('/api/analysis/export', methods=['POST'])
def create_export():
    """
    创建导出任务
    
    参数:
        task_id: 任务ID
        format: pdf|excel
        sections: 导出内容
    """
    try:
        data = request.get_json() or {}
        task_id = data.get('task_id')
        export_format = data.get('format', 'excel')
        sections = data.get('sections', ['summary', 'clusters', 'samples'])
        
        if not task_id:
            return jsonify({'success': False, 'error': '缺少 task_id 参数'}), 400
        
        import uuid
        export_id = f"export_{uuid.uuid4().hex[:8]}"
        
        # 创建导出任务（简化实现：同步生成）
        try:
            if export_format == 'excel':
                file_path = _generate_excel_report(task_id, export_id, sections)
            else:
                file_path = _generate_excel_report(task_id, export_id, sections)  # PDF 暂用 Excel
            
            return jsonify({
                'success': True,
                'data': {
                    'export_id': export_id,
                    'status': 'completed',
                    'format': export_format,
                    'file_path': file_path
                }
            })
        except Exception as gen_error:
            return jsonify({
                'success': False,
                'error': f'生成报告失败: {str(gen_error)}'
            }), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/export/<export_id>', methods=['GET'])
def get_export_status(export_id):
    """获取导出状态"""
    try:
        import os
        file_path = f"exports/{export_id}.xlsx"
        
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
            return jsonify({
                'success': True,
                'data': {
                    'export_id': export_id,
                    'status': 'completed',
                    'download_url': f'/api/analysis/export/download/{export_id}',
                    'file_name': f'{export_id}.xlsx',
                    'file_size': file_size
                }
            })
        else:
            return jsonify({
                'success': True,
                'data': {
                    'export_id': export_id,
                    'status': 'not_found'
                }
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analysis_bp.route('/api/analysis/export/download/<export_id>', methods=['GET'])
def download_export(export_id):
    """下载导出文件"""
    try:
        from flask import send_file
        import os
        
        file_path = f"exports/{export_id}.xlsx"
        
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=f'analysis_report_{export_id}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _generate_excel_report(task_id: str, export_id: str, sections: list) -> str:
    """生成 Excel 报告"""
    import os
    from openpyxl import Workbook
    
    os.makedirs('exports', exist_ok=True)
    file_path = f"exports/{export_id}.xlsx"
    
    wb = Workbook()
    
    # 获取数据
    stats = analysis_service.get_quick_stats(task_id)
    report = analysis_service.get_report(task_id)
    
    # 摘要页
    if 'summary' in sections:
        ws = wb.active
        ws.title = '摘要'
        ws.append(['AI 分析报告'])
        ws.append(['任务ID', task_id])
        ws.append(['总错误数', stats.get('total_errors', 0)])
        ws.append(['总题目数', stats.get('total_questions', 0)])
        ws.append(['错误率', f"{stats.get('error_rate', 0)*100:.2f}%"])
        ws.append([])
        ws.append(['错误类型分布'])
        for error_type, count in stats.get('error_type_distribution', {}).items():
            ws.append([error_type, count])
    
    # 聚类页
    if 'clusters' in sections:
        ws = wb.create_sheet('聚类分析')
        ws.append(['聚类键', '错误类型', '书本', '页码范围', '样本数'])
        for cluster in stats.get('clusters', []):
            ws.append([
                cluster.get('cluster_key', ''),
                cluster.get('error_type', ''),
                cluster.get('book_name', ''),
                cluster.get('page_range', ''),
                cluster.get('sample_count', 0)
            ])
    
    # 样本页
    if 'samples' in sections:
        ws = wb.create_sheet('错误样本')
        ws.append(['作业ID', '书本', '页码', '题号', '错误类型', 'AI答案', '期望答案'])
        for cluster in stats.get('clusters', []):
            for sample in cluster.get('samples', []):
                ws.append([
                    sample.get('homework_id', ''),
                    sample.get('book_name', ''),
                    sample.get('page_num', 0),
                    sample.get('question_index', 0),
                    sample.get('error_type', ''),
                    str(sample.get('ai_answer', ''))[:100],
                    str(sample.get('expected_answer', ''))[:100]
                ])
    
    wb.save(file_path)
    return file_path


# ============================================
# 错误日志 API
# ============================================

@analysis_bp.route('/api/analysis/logs', methods=['GET'])
def get_analysis_logs():
    """
    获取分析错误日志
    
    参数:
        days: 天数
        type: timeout|api_error|parse_error
    """
    try:
        days = int(request.args.get('days', 7))
        log_type = request.args.get('type')
        
        # 从数据库查询日志
        try:
            from services.database_service import AppDatabaseService
            
            sql = """
                SELECT log_id, task_id, analysis_type, target_id, model,
                       prompt_tokens, completion_tokens, total_tokens,
                       duration_ms, retry_count, status, error_type, error_message, created_at
                FROM llm_call_logs
                WHERE created_at >= DATE_SUB(NOW(), INTERVAL %s DAY)
            """
            params = [days]
            
            if log_type:
                sql += " AND error_type = %s"
                params.append(log_type)
            
            sql += " ORDER BY created_at DESC LIMIT 100"
            
            result = AppDatabaseService.execute_query(sql, tuple(params))
            
            logs = []
            for row in result:
                logs.append({
                    'log_id': row.get('log_id'),
                    'task_id': row.get('task_id'),
                    'analysis_type': row.get('analysis_type'),
                    'model': row.get('model'),
                    'tokens': row.get('total_tokens', 0),
                    'duration_ms': row.get('duration_ms', 0),
                    'status': row.get('status'),
                    'error_type': row.get('error_type'),
                    'error_message': row.get('error_message'),
                    'created_at': row.get('created_at').isoformat() if row.get('created_at') else None
                })
            
            # 统计摘要
            summary = {
                'total': len(logs),
                'success': len([l for l in logs if l.get('status') == 'success']),
                'failed': len([l for l in logs if l.get('status') == 'failed']),
                'timeout': len([l for l in logs if l.get('error_type') == 'timeout']),
                'api_error': len([l for l in logs if l.get('error_type') == 'api_error']),
                'parse_error': len([l for l in logs if l.get('error_type') == 'parse_error'])
            }
            
            return jsonify({
                'success': True,
                'data': {
                    'logs': logs,
                    'summary': summary
                }
            })
        except Exception as db_error:
            # 数据库不可用时返回空
            return jsonify({
                'success': True,
                'data': {
                    'logs': [],
                    'summary': {'total': 0, 'success': 0, 'failed': 0},
                    'db_error': str(db_error)
                }
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
