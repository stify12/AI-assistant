"""
任务整合评估路由模块
提供多任务数据整合分析功能，重点关注批改准确率(correct字段)
"""
import os
import io
import json
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, send_file

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from services.storage_service import StorageService

task_consolidation_bp = Blueprint('task_consolidation', __name__)

# 学科ID映射
SUBJECT_MAP = {
    0: '英语',
    1: '语文',
    2: '数学',
    3: '物理',
    4: '化学',
    5: '生物',
    6: '地理'
}


@task_consolidation_bp.route('/task-consolidation')
def task_consolidation_page():
    """任务整合评估页面"""
    return render_template('task-consolidation.html')


@task_consolidation_bp.route('/api/consolidation/tasks', methods=['GET'])
def get_tasks_for_consolidation():
    """
    获取可用于整合的任务列表
    支持按学科筛选，只返回已完成的任务
    """
    subject_id = request.args.get('subject_id', type=int)
    
    tasks = []
    for filename in StorageService.list_batch_tasks():
        task_id = filename.replace('.json', '')
        task_data = StorageService.load_batch_task(task_id)
        if not task_data:
            continue
        
        # 只返回已完成的任务
        if task_data.get('status') != 'completed':
            continue
        
        # 按学科筛选
        task_subject_id = task_data.get('subject_id')
        if subject_id is not None and task_subject_id != subject_id:
            continue
        
        homework_items = task_data.get('homework_items', [])
        overall_report = task_data.get('overall_report', {})
        
        # 提取书本和页码信息
        book_name = ''
        page_nums = set()
        for hw in homework_items:
            if not book_name and hw.get('book_name'):
                book_name = hw.get('book_name', '')
            if hw.get('page_num'):
                page_nums.add(hw.get('page_num'))
        
        page_range = ''
        if page_nums:
            sorted_pages = sorted(page_nums)
            if len(sorted_pages) == 1:
                page_range = f"P{sorted_pages[0]}"
            else:
                page_range = f"P{sorted_pages[0]}-{sorted_pages[-1]}"
        
        tasks.append({
            'task_id': task_data.get('task_id', task_id),
            'name': task_data.get('name', ''),
            'subject_id': task_subject_id,
            'subject_name': SUBJECT_MAP.get(task_subject_id, '未知'),
            'test_condition_name': task_data.get('test_condition_name', ''),
            'created_at': task_data.get('created_at', ''),
            'homework_count': len(homework_items),
            'total_questions': overall_report.get('total_questions', 0),
            'correct_questions': overall_report.get('correct_questions', 0),
            'overall_accuracy': overall_report.get('overall_accuracy', 0),
            'book_name': book_name,
            'page_range': page_range
        })
    
    # 按创建时间倒序
    tasks.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    return jsonify({'success': True, 'tasks': tasks})


@task_consolidation_bp.route('/api/consolidation/analyze', methods=['POST'])
def analyze_consolidated_tasks():
    """
    整合分析多个任务的评估数据
    重点分析批改字段correct的准确率
    """
    try:
        data = request.get_json() or {}
        task_ids = data.get('task_ids', [])
        
        if not task_ids or len(task_ids) < 1:
            return jsonify({'success': False, 'error': '请至少选择一个任务'})
        
        # 加载所有任务数据
        all_tasks_data = []
        for task_id in task_ids:
            task_data = StorageService.load_batch_task(task_id)
            if task_data:
                all_tasks_data.append(task_data)
        
        if not all_tasks_data:
            return jsonify({'success': False, 'error': '无法加载任务数据'})
        
        # 整合分析
        result = consolidate_analysis(all_tasks_data)
        
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@task_consolidation_bp.route('/api/consolidation/export', methods=['POST'])
def export_accuracy_analysis():
    """
    导出批改准确率分析表格为Excel
    将所有选择的批量评估作业数据整合为一份导出
    格式：题型、总份数、题型总数、错误数量、批改准确率、综合准确率
    """
    data = request.get_json()
    task_ids = data.get('task_ids', [])
    
    if not task_ids or len(task_ids) < 1:
        return jsonify({'success': False, 'error': '请至少选择一个任务'})
    
    # 加载所有任务数据
    all_tasks_data = []
    for task_id in task_ids:
        task_data = StorageService.load_batch_task(task_id)
        if task_data:
            all_tasks_data.append(task_data)
    
    if not all_tasks_data:
        return jsonify({'success': False, 'error': '无法加载任务数据'})
    
    # 整合分析
    result = consolidate_analysis(all_tasks_data)
    summary = result.get('summary', {})
    type_accuracy = result.get('type_accuracy', [])
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "批改准确率分析"
    
    # 样式定义
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True)
    cell_font = Font(name='微软雅黑', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    
    # 标题行
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = 'AI批改准确率整合分析'
    title_cell.font = title_font
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 35
    
    # 汇总信息行
    ws.merge_cells('A2:F2')
    summary_cell = ws['A2']
    summary_cell.value = f"共整合 {summary.get('total_tasks', 0)} 个任务，{summary.get('total_homework', 0)} 份作业，{summary.get('total_questions', 0)} 道题目"
    summary_cell.font = Font(name='微软雅黑', size=11)
    summary_cell.alignment = center_align
    ws.row_dimensions[2].height = 25
    
    # 表头
    headers = ['题型', '总份数', '题型总数', '错误数量', '批改准确率', '综合准确率']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
    ws.row_dimensions[3].height = 30
    
    # 写入整合后的题型数据
    current_row = 4
    type_count = len(type_accuracy)
    overall_accuracy = summary.get('overall_accuracy', 0)
    overall_acc_str = f"{overall_accuracy * 100:.2f}%"
    
    for i, type_item in enumerate(type_accuracy):
        row = current_row + i
        
        # 题型数据
        ws.cell(row=row, column=1, value=type_item['name']).font = cell_font
        ws.cell(row=row, column=2, value=summary.get('total_homework', 0)).font = cell_font
        ws.cell(row=row, column=3, value=type_item['total']).font = cell_font
        ws.cell(row=row, column=4, value=type_item['total'] - type_item['match']).font = cell_font
        ws.cell(row=row, column=5, value=f"{type_item['accuracy'] * 100:.2f}%").font = cell_font
        
        # 设置单元格样式
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.border = thin_border
        
        ws.row_dimensions[row].height = 25
    
    # 综合准确率列（合并单元格）
    if type_count > 1:
        ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row + type_count - 1, end_column=6)
    
    ws.cell(row=current_row, column=6, value=overall_acc_str).font = cell_font
    for row in range(current_row, current_row + type_count):
        cell = ws.cell(row=row, column=6)
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row += type_count + 1
    
    # 添加汇总行
    ws.cell(row=current_row, column=1, value='合计').font = Font(name='微软雅黑', size=10, bold=True)
    ws.cell(row=current_row, column=2, value=summary.get('total_homework', 0)).font = Font(name='微软雅黑', size=10, bold=True)
    ws.cell(row=current_row, column=3, value=summary.get('total_questions', 0)).font = Font(name='微软雅黑', size=10, bold=True)
    ws.cell(row=current_row, column=4, value=summary.get('grading_distribution', {}).get('mismatch', 0)).font = Font(name='微软雅黑', size=10, bold=True)
    ws.cell(row=current_row, column=5, value=overall_acc_str).font = Font(name='微软雅黑', size=10, bold=True)
    ws.cell(row=current_row, column=6, value=overall_acc_str).font = Font(name='微软雅黑', size=10, bold=True)
    
    for col in range(1, 7):
        cell = ws.cell(row=current_row, column=col)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    ws.row_dimensions[current_row].height = 28
    
    # 设置列宽
    column_widths = [12, 10, 12, 12, 14, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ============ 创建详细Sheet：每个任务的单独统计 ============
    ws_detail = wb.create_sheet(title="任务详细统计")
    
    # 详细Sheet样式定义
    detail_title_font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')
    detail_header_font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79')
    detail_cell_font = Font(name='微软雅黑', size=10)
    detail_header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    detail_center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    detail_border = Border(
        left=Side(style='thin', color='808080'),
        right=Side(style='thin', color='808080'),
        top=Side(style='thin', color='808080'),
        bottom=Side(style='thin', color='808080')
    )
    
    # 详细Sheet标题
    detail_headers = ['书籍名称', '页码', '测试时间', '科目', '题型', '份数', '题型总数', '错误数量', '批改准确率', '综合准确率']
    ws_detail.merge_cells('A1:J1')
    detail_title_cell = ws_detail['A1']
    
    # 获取学科名称用于标题
    subject_name = '数学'
    if all_tasks_data:
        subject_id = all_tasks_data[0].get('subject_id', 2)
        subject_name = SUBJECT_MAP.get(subject_id, '数学')
    detail_title_cell.value = f'{subject_name} AI 批改准确率分析'
    detail_title_cell.font = detail_title_font
    detail_title_cell.alignment = detail_center_align
    ws_detail.row_dimensions[1].height = 30
    
    # 详细Sheet表头
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=2, column=col, value=header)
        cell.font = detail_header_font
        cell.alignment = detail_center_align
        cell.border = detail_border
        cell.fill = detail_header_fill
    ws_detail.row_dimensions[2].height = 28
    
    # 按任务分组写入详细数据
    task_type_analysis = result.get('task_type_analysis', [])
    
    # 按任务ID分组
    task_groups = {}
    for item in task_type_analysis:
        task_id = item['task_id']
        if task_id not in task_groups:
            task_groups[task_id] = {
                'task_name': item['task_name'],
                'page_range': item['page_range'],
                'created_at': item['created_at'],
                'subject_name': item['subject_name'],
                'homework_count': item['homework_count'],
                'overall_accuracy': item['overall_accuracy'],
                'types': []
            }
        task_groups[task_id]['types'].append({
            'type_name': item['type_name'],
            'type_key': item['type_key'],
            'homework_count': item['homework_count'],
            'type_total': item['type_total'],
            'type_error': item['type_error'],
            'type_accuracy': item['type_accuracy']
        })
    
    # 写入每个任务的数据
    detail_row = 3
    type_order = ['choice', 'objective_fill', 'subjective']
    type_name_map = {'choice': '选择题', 'objective_fill': '填空题', 'subjective': '主观题'}
    
    for task_id, task_info in task_groups.items():
        # 获取该任务的书籍名称
        book_name = task_info['task_name']
        # 从任务数据中获取更准确的书籍名称
        for task_data in all_tasks_data:
            if task_data.get('task_id') == task_id:
                hw_items = task_data.get('homework_items', [])
                for hw in hw_items:
                    if hw.get('book_name'):
                        book_name = hw.get('book_name')
                        break
                break
        
        # 格式化测试时间
        test_date = ''
        if task_info['created_at']:
            try:
                dt = datetime.fromisoformat(task_info['created_at'].replace('Z', '+00:00'))
                test_date = dt.strftime('%Y/%m/%d')
            except:
                test_date = task_info['created_at'][:10].replace('-', '/')
        
        # 按题型顺序排列
        types_data = task_info['types']
        sorted_types = []
        for type_key in type_order:
            for t in types_data:
                if t['type_key'] == type_key:
                    sorted_types.append(t)
                    break
        
        # 如果没有某种题型的数据，添加空行
        if not sorted_types:
            for type_key in type_order:
                sorted_types.append({
                    'type_name': type_name_map[type_key],
                    'type_key': type_key,
                    'homework_count': task_info['homework_count'],
                    'type_total': 0,
                    'type_error': 0,
                    'type_accuracy': 0
                })
        
        row_count = len(sorted_types)
        start_row = detail_row
        end_row = detail_row + row_count - 1
        
        # 写入每个题型行
        for i, type_data in enumerate(sorted_types):
            row = detail_row + i
            
            # 题型数据
            ws_detail.cell(row=row, column=5, value=type_data['type_name']).font = detail_cell_font
            ws_detail.cell(row=row, column=6, value=type_data['homework_count']).font = detail_cell_font
            ws_detail.cell(row=row, column=7, value=type_data['type_total']).font = detail_cell_font
            ws_detail.cell(row=row, column=8, value=type_data['type_error']).font = detail_cell_font
            
            # 批改准确率
            acc_str = f"{type_data['type_accuracy'] * 100:.2f}%" if type_data['type_total'] > 0 else '-'
            ws_detail.cell(row=row, column=9, value=acc_str).font = detail_cell_font
            
            # 设置单元格样式
            for col in range(5, 10):
                cell = ws_detail.cell(row=row, column=col)
                cell.alignment = detail_center_align
                cell.border = detail_border
            
            ws_detail.row_dimensions[row].height = 22
        
        # 合并单元格：书籍名称、页码、测试时间、科目、综合准确率
        if row_count > 1:
            ws_detail.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws_detail.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws_detail.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
            ws_detail.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
            ws_detail.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
        
        # 写入合并单元格的值
        ws_detail.cell(row=start_row, column=1, value=book_name).font = detail_cell_font
        ws_detail.cell(row=start_row, column=2, value=task_info['page_range']).font = detail_cell_font
        ws_detail.cell(row=start_row, column=3, value=test_date).font = detail_cell_font
        ws_detail.cell(row=start_row, column=4, value=task_info['subject_name']).font = detail_cell_font
        
        overall_acc_str = f"{task_info['overall_accuracy'] * 100:.2f}%"
        ws_detail.cell(row=start_row, column=10, value=overall_acc_str).font = detail_cell_font
        
        # 设置合并单元格的边框和对齐
        for row in range(start_row, end_row + 1):
            for col in [1, 2, 3, 4, 10]:
                cell = ws_detail.cell(row=row, column=col)
                cell.alignment = detail_center_align
                cell.border = detail_border
        
        detail_row = end_row + 1
    
    # 设置详细Sheet列宽
    detail_column_widths = [14, 8, 12, 8, 12, 8, 10, 10, 12, 12]
    for i, width in enumerate(detail_column_widths, 1):
        ws_detail.column_dimensions[get_column_letter(i)].width = width
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"批改准确率整合分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def consolidate_analysis(tasks_data):
    """
    整合多个任务的分析数据
    
    批改准确率的定义：基准答案的correct字段与AI批改的correct字段一致的比例
    
    Args:
        tasks_data: 任务数据列表
        
    Returns:
        整合分析结果
    """
    # 总体统计
    total_tasks = len(tasks_data)
    total_homework = 0
    total_questions = 0  # 有基准数据的题目总数
    total_grading_match = 0  # 批改一致的题目数
    
    # 按任务统计
    task_stats = []
    
    # 批改结果分布统计
    grading_match_count = 0    # 基准correct与AI correct一致
    grading_mismatch_count = 0 # 基准correct与AI correct不一致
    grading_unknown_count = 0  # 缺少基准或AI的correct数据
    
    # 错误类型统计
    error_type_stats = {}
    
    # 题目级别统计（按题号聚合）
    question_stats = {}  # {index: {total: n, match: n, mismatch: n, errors: []}}

    # 按页码统计
    page_stats = {}  # {page_num: {total: n, match: n}}

    # 按题型统计（三类：选择题、填空题、主观题）
    type_stats = {
        'choice': {'name': '选择题', 'total': 0, 'match': 0},
        'objective_fill': {'name': '填空题', 'total': 0, 'match': 0},
        'subjective': {'name': '主观题', 'total': 0, 'match': 0}
    }

    # 按作业统计
    homework_details = []

    # 按任务分组的题型统计（用于批改准确率分析表格）
    task_type_analysis = []

    for task_data in tasks_data:
        task_id = task_data.get('task_id', '')
        task_name = task_data.get('name', '')
        homework_items = task_data.get('homework_items', [])
        overall_report = task_data.get('overall_report', {})
        created_at = task_data.get('created_at', '')
        subject_id = task_data.get('subject_id', 2)

        task_total_questions = 0
        task_grading_match = 0
        task_homework_count = len(homework_items)

        # 任务级别的题型统计（三类：选择题、填空题、主观题）
        task_type_stats = {
            'choice': {'name': '选择题', 'total': 0, 'match': 0},
            'objective_fill': {'name': '填空题', 'total': 0, 'match': 0},
            'subjective': {'name': '主观题', 'total': 0, 'match': 0}
        }

        # 收集页码范围
        page_nums = set()

        total_homework += task_homework_count

        for hw in homework_items:
            hw_id = hw.get('homework_id', '')
            page_num = hw.get('page_num', 0)
            evaluation = hw.get('evaluation', {})
            
            # 收集页码
            page_nums.add(page_num)
            
            # 直接从evaluation中获取题型统计（与批量评估保持一致）
            by_question_type = evaluation.get('by_question_type', {})
            
            hw_total = evaluation.get('total_questions', 0)
            hw_correct = evaluation.get('correct_count', 0)
            hw_error = evaluation.get('error_count', 0)
            
            # 累加题型统计（使用批量评估已计算的数据，三类：选择题/填空题/主观题）
            for qtype in ['choice', 'objective_fill', 'subjective']:
                qtype_data = by_question_type.get(qtype, {})
                qtype_total = qtype_data.get('total', 0)
                # 使用grading_correct作为批改一致数（如果有），否则使用correct
                qtype_correct = qtype_data.get('grading_correct', qtype_data.get('correct', 0))
                
                if qtype_total > 0:
                    type_stats[qtype]['total'] += qtype_total
                    type_stats[qtype]['match'] += qtype_correct
                    task_type_stats[qtype]['total'] += qtype_total
                    task_type_stats[qtype]['match'] += qtype_correct
            
            # 累加总体统计
            total_questions += hw_total
            task_total_questions += hw_total
            total_grading_match += hw_correct
            task_grading_match += hw_correct
            grading_match_count += hw_correct
            grading_mismatch_count += hw_error
            
            # 初始化页码统计
            if page_num not in page_stats:
                page_stats[page_num] = {'total': 0, 'match': 0}
            page_stats[page_num]['total'] += hw_total
            page_stats[page_num]['match'] += hw_correct
            
            # 从evaluation中获取错误类型和错误详情
            if evaluation:
                errors = evaluation.get('errors', [])
                for err in errors:
                    err_type = err.get('error_type', '未知错误')
                    error_type_stats[err_type] = error_type_stats.get(err_type, 0) + 1
                    
                    # 记录题目级别统计
                    index = str(err.get('index', ''))
                    if index:
                        if index not in question_stats:
                            question_stats[index] = {
                                'total': 0,
                                'match': 0,
                                'mismatch': 0,
                                'errors': []
                            }
                        question_stats[index]['mismatch'] += 1
                        question_stats[index]['errors'].append({
                            'task_id': task_id,
                            'task_name': task_name,
                            'homework_id': hw_id,
                            'page_num': page_num,
                            'index': index,
                            'error_type': err_type
                        })
            
            # 作业详情
            homework_details.append({
                'task_id': task_id,
                'task_name': task_name,
                'homework_id': hw_id,
                'student_name': hw.get('student_name', ''),
                'page_num': page_num,
                'total_questions': hw_total,
                'match_count': hw_correct,
                'accuracy': hw_correct / hw_total if hw_total > 0 else 0,
                'errors': evaluation.get('errors', [])
            })
        
        # 任务级别统计
        task_accuracy = task_grading_match / task_total_questions if task_total_questions > 0 else 0
        task_stats.append({
            'task_id': task_id,
            'name': task_name,
            'subject_name': SUBJECT_MAP.get(task_data.get('subject_id'), '未知'),
            'test_condition_name': task_data.get('test_condition_name', ''),
            'homework_count': task_homework_count,
            'total_questions': task_total_questions,
            'match_count': task_grading_match,
            'accuracy': task_accuracy,
            'created_at': task_data.get('created_at', '')
        })
        
        # 生成页码范围字符串
        sorted_pages = sorted(page_nums)
        if sorted_pages:
            page_range = f"{min(sorted_pages)}-{max(sorted_pages)}" if len(sorted_pages) > 1 else str(sorted_pages[0])
        else:
            page_range = ''
        
        # 任务级别的题型分析（用于批改准确率分析表格）
        for type_key, stats in task_type_stats.items():
            if stats['total'] > 0:
                type_accuracy = stats['match'] / stats['total']
                task_type_analysis.append({
                    'task_id': task_id,
                    'task_name': task_name,
                    'page_range': page_range,
                    'created_at': created_at,
                    'subject_name': SUBJECT_MAP.get(subject_id, '未知'),
                    'type_name': stats['name'],
                    'type_key': type_key,
                    'homework_count': task_homework_count,
                    'type_total': stats['total'],
                    'type_error': stats['total'] - stats['match'],
                    'type_accuracy': type_accuracy,
                    'overall_accuracy': task_accuracy
                })
    
    # 计算总体批改准确率
    overall_accuracy = total_grading_match / total_questions if total_questions > 0 else 0
    
    # 高频不一致题目（按不一致次数排序）
    high_error_questions = []
    for index, stats in question_stats.items():
        if stats['mismatch'] > 0:
            high_error_questions.append({
                'index': index,
                'total': stats['total'],
                'match': stats['match'],
                'mismatch': stats['mismatch'],
                'error_rate': stats['mismatch'] / stats['total'] if stats['total'] > 0 else 0,
                'sample_errors': stats['errors'][:3]
            })
    high_error_questions.sort(key=lambda x: x['mismatch'], reverse=True)
    
    # 页码准确率统计
    page_accuracy_list = []
    for page_num, stats in page_stats.items():
        page_accuracy_list.append({
            'page_num': page_num,
            'total': stats['total'],
            'match': stats['match'],
            'accuracy': stats['match'] / stats['total'] if stats['total'] > 0 else 0
        })
    page_accuracy_list.sort(key=lambda x: x['page_num'])
    
    # 错误类型分布
    error_type_list = [
        {'type': k, 'count': v}
        for k, v in sorted(error_type_stats.items(), key=lambda x: x[1], reverse=True)
    ]
    
    # 题型准确率列表
    type_accuracy_list = []
    for type_key, stats in type_stats.items():
        if stats['total'] > 0:
            type_accuracy_list.append({
                'type_key': type_key,
                'name': stats['name'],
                'total': stats['total'],
                'match': stats['match'],
                'accuracy': stats['match'] / stats['total']
            })
    # 按total降序排列
    type_accuracy_list.sort(key=lambda x: x['total'], reverse=True)
    
    return {
        'summary': {
            'total_tasks': total_tasks,
            'total_homework': total_homework,
            'total_questions': total_questions,
            'total_grading_match': total_grading_match,
            'overall_accuracy': overall_accuracy,
            'grading_distribution': {
                'match': grading_match_count,
                'mismatch': grading_mismatch_count,
                'unknown': grading_unknown_count
            }
        },
        'task_stats': task_stats,
        'high_error_questions': high_error_questions[:20],
        'page_accuracy': page_accuracy_list,
        'type_accuracy': type_accuracy_list,
        'task_type_analysis': task_type_analysis,
        'error_types': error_type_list,
        'homework_details': homework_details
    }


def flatten_homework_result(homework_result):
    """
    展开 homework_result 中的嵌套 children 结构为扁平数组
    """
    if not homework_result:
        return []
    
    flattened = []
    for item in homework_result:
        children = item.get('children', [])
        if children and len(children) > 0:
            for child in children:
                flattened.append(child)
        else:
            flattened.append(item)
    return flattened


def classify_question_type_key(question_data):
    """
    根据题目数据判断题目类型（三类互不包含，与批量评估保持一致）
    
    分类规则：
    1. 选择题: bvalue=1(单选)、2(多选)、3(判断)
    2. 填空题: bvalue='4'（不区分客观/主观填空，统一归为填空题）
    3. 主观题: 其他所有
    
    Args:
        question_data: 题目数据
        
    Returns:
        'choice' | 'objective_fill' | 'subjective'
    """
    if not question_data:
        return 'subjective'
    
    bvalue = str(question_data.get('bvalue', ''))
    
    # 选择题: bvalue=1(单选)、2(多选)、3(判断)
    if bvalue in ('1', '2', '3'):
        return 'choice'
    
    # 填空题: bvalue='4'（客观填空和主观填空统一归为填空题）
    if bvalue == '4':
        return 'objective_fill'
    
    # 主观题: 其他所有
    return 'subjective'


def get_correct_value(item):
    """
    获取判断结果字段值，兼容多种格式
    统一返回 "yes" 或 "no" 字符串
    """
    if not item:
        return ''
    
    # 优先检查 correct 字段
    if 'correct' in item:
        val = item['correct']
        if isinstance(val, bool):
            return 'yes' if val else 'no'
        if isinstance(val, str):
            val_lower = val.lower().strip()
            if val_lower in ('yes', 'true', '1'):
                return 'yes'
            elif val_lower in ('no', 'false', '0'):
                return 'no'
            return val_lower
    
    # 检查 isRight 字段
    if 'isRight' in item:
        val = item['isRight']
        if isinstance(val, bool):
            return 'yes' if val else 'no'
        if isinstance(val, str):
            return 'yes' if val.lower() in ('true', 'yes', '1') else 'no'
    
    # 检查 isCorrect 字段
    if 'isCorrect' in item:
        val = item['isCorrect']
        if isinstance(val, bool):
            return 'yes' if val else 'no'
        if isinstance(val, str):
            return 'yes' if val.lower() in ('true', 'yes', '1') else 'no'
    
    return ''
