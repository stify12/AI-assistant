"""
异常检测服务模块 (US-26)

提供准确率异常自动检测和告警功能。
支持基于统计学方法的异常检测，自动记录异常日志。

任务级题目异常检测：
- 全员错误：某题所有学生的AI批改结果都与基准不一致（可能是数据集标注错误）
- 高异常：有人批改对的情况下，错误类型是"识别正确-判断错误"（AI识别准确但判断错误）
- 低异常：有人批改对的情况下，其他错误类型（识别错误等）

一致性计算：
- 一致性 = 正常题数 / (总题数 - 全员错误题数) × 100%
- 排除全员错误的影响，因为可能是数据集标注问题
"""
import uuid
import json
import os
import re
import statistics
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from collections import defaultdict

from .database_service import AppDatabaseService
from .storage_service import StorageService


class AnomalyService:
    """异常检测服务类"""
    
    # 默认阈值：2个标准差
    DEFAULT_THRESHOLD_SIGMA = 2.0
    
    # 配置缓存
    _config = {
        'threshold_sigma': 2.0,
        'min_samples': 5,  # 最少需要5个历史样本才能检测
    }
    
    # 高异常错误类型（识别正确但判断错误）
    HIGH_ANOMALY_ERROR_TYPES = ['识别正确-判断错误']
    
    # ========== 任务级题目异常检测 ==========
    
    @staticmethod
    def _collect_question_indices(data_value: list, result: set = None) -> set:
        """从 data_value 递归收集所有子题的题号（只收集叶子节点）"""
        if result is None:
            result = set()
        for item in data_value:
            children = item.get('children', [])
            if children:
                # 有子题，递归处理子题
                AnomalyService._collect_question_indices(children, result)
            else:
                # 叶子节点，收集题号
                idx = item.get('index')
                if idx:
                    result.add(str(idx))
        return result
    
    @staticmethod
    def detect_question_anomalies(task_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        检测任务中的题目异常
        
        异常类型：
        1. 全员错误：同一页码的某题，所有包含该页的作业都判错
        2. 高异常：有人批改对的情况下，错误类型是"识别正确-判断错误"
        3. 低异常：有人批改对的情况下，其他错误类型
        4. 正常：所有人都做对的题目
        
        Args:
            task_data: 完整的任务数据
            
        Returns:
            {
                'summary': {...},
                'anomalies': [...]
            }
        """
        homework_items = task_data.get('homework_items', [])
        completed_items = [h for h in homework_items if h.get('status') == 'completed' and h.get('evaluation')]
        
        if not completed_items:
            return {
                'summary': {
                    'universal_errors': 0,
                    'high_anomaly': 0,
                    'low_anomaly': 0,
                    'normal': 0,
                    'total_questions': 0,
                    'anomaly_rate': 0
                },
                'anomalies': []
            }
        
        # 按页码统计作业数
        page_homework_count = defaultdict(int)
        for item in completed_items:
            page_num = item.get('page_num', '?')
            page_homework_count[page_num] += 1
        
        # 按题目聚合统计 (page_num + question_index)
        # 使用 evaluation 中的题目信息作为统计依据
        question_stats = defaultdict(lambda: {
            'total': 0,
            'correct': 0,
            'error': 0,
            'samples': [],
            'error_types': defaultdict(int),
            'error_samples': [],
            'page_num': '?',
            'question_index': '?'
        })
        
        # 遍历每个作业，统计正确和错误
        for item in completed_items:
            page_num = item.get('page_num', '?')
            evaluation = item.get('evaluation', {})
            student_id = item.get('student_id', item.get('homework_id', ''))
            student_name = item.get('student_name', '')
            
            total_questions = evaluation.get('total_questions', 0)
            correct_count = evaluation.get('correct_count', 0)
            errors = evaluation.get('errors', [])
            
            # 解析 homework_result 获取 AI 识别的答案（用于正确题目）
            homework_result_str = item.get('homework_result', '[]')
            try:
                homework_result = json.loads(homework_result_str) if isinstance(homework_result_str, str) else homework_result_str
            except:
                homework_result = []
            
            # 构建 homework_result 的索引映射 (index -> item)，支持多种格式
            hw_result_map = {}
            def build_hw_map(items):
                for hw_item in items:
                    idx = hw_item.get('index')
                    temp_idx = hw_item.get('tempIndex')
                    if idx:
                        hw_result_map[str(idx)] = hw_item
                        # 也用纯数字索引
                        import re
                        nums = re.findall(r'\d+', str(idx))
                        if nums:
                            hw_result_map[nums[0]] = hw_item
                    if temp_idx is not None:
                        hw_result_map[f'temp_{temp_idx}'] = hw_item
                    children = hw_item.get('children', [])
                    if children:
                        build_hw_map(children)
            build_hw_map(homework_result)
            
            # 获取基准效果（从匹配的数据集中加载）
            base_effect_map = {}
            matched_dataset_id = item.get('matched_dataset')
            if matched_dataset_id:
                ds_data = StorageService.load_dataset(matched_dataset_id)
                if ds_data:
                    page_key = str(page_num)
                    base_effects = ds_data.get('base_effects', {}).get(page_key, [])
                    for be in base_effects:
                        be_idx = be.get('index')
                        be_temp_idx = be.get('tempIndex')
                        if be_idx:
                            base_effect_map[str(be_idx)] = be
                        if be_temp_idx is not None:
                            base_effect_map[f'temp_{be_temp_idx}'] = be
            
            # 收集错误题目的索引
            error_indices = set()
            for err in errors:
                q_index = str(err.get('index', '?'))
                error_indices.add(q_index)
                
                q_key = f"{page_num}_{q_index}"
                stats = question_stats[q_key]
                stats['page_num'] = page_num
                stats['question_index'] = q_index
                stats['total'] += 1
                stats['error'] += 1
                
                error_type = err.get('error_type', '未知错误')
                stats['error_types'][error_type] += 1
                
                base_effect = err.get('base_effect', {})
                ai_result = err.get('ai_result', {})
                
                error_sample = {
                    'student_id': student_id,
                    'student_name': student_name,
                    'result': 'error',
                    'hw_answer': ai_result.get('userAnswer', ''),
                    'base_answer': base_effect.get('answer', ''),
                    'base_user': base_effect.get('userAnswer', ''),
                    'error_type': error_type
                }
                stats['error_samples'].append(error_sample)
                if len(stats['samples']) < 10:
                    stats['samples'].append(error_sample)
            
            # 处理正确的题目：从基准效果中获取所有题目，排除错误的
            for be_idx, be in base_effect_map.items():
                if be_idx.startswith('temp_'):
                    continue  # 跳过 tempIndex 索引，避免重复
                
                if be_idx in error_indices:
                    continue  # 已经在错误中处理过了
                
                q_key = f"{page_num}_{be_idx}"
                stats = question_stats[q_key]
                stats['page_num'] = page_num
                stats['question_index'] = be_idx
                stats['total'] += 1
                stats['correct'] += 1
                
                # 获取 AI 识别的答案
                hw_item = hw_result_map.get(be_idx, {})
                if not hw_item:
                    # 尝试用 tempIndex 查找
                    temp_idx = be.get('tempIndex')
                    if temp_idx is not None:
                        hw_item = hw_result_map.get(f'temp_{temp_idx}', {})
                
                hw_answer = hw_item.get('userAnswer', '')
                base_answer = be.get('answer', '')
                base_user = be.get('userAnswer', '')
                
                correct_sample = {
                    'student_id': student_id,
                    'student_name': student_name,
                    'result': 'correct',
                    'hw_answer': hw_answer,
                    'base_answer': base_answer,
                    'base_user': base_user
                }
                if len(stats['samples']) < 10:
                    stats['samples'].append(correct_sample)
        
        # 分析异常
        anomalies = []
        universal_errors = 0
        high_anomaly = 0
        low_anomaly = 0
        normal = 0
        
        for q_key, stats in question_stats.items():
            if stats['total'] == 0:
                continue
            
            error_rate = stats['error'] / stats['total']
            page_num = stats.get('page_num', '?')
            q_index = stats.get('question_index', '?')
            
            # 获取该页码的作业总数
            page_total = page_homework_count.get(page_num, 0)
            
            anomaly = None
            
            # 1. 全员错误检测（新定义）
            # 条件：该题目所有学生的AI批改结果都与基准不一致（全部异常）
            # 即：没有任何一个学生该题是正常的
            is_universal_error = (
                error_rate == 1.0 and 
                stats['total'] >= 2 and 
                stats['correct'] == 0  # 没有任何正确的
            )
            
            if is_universal_error:
                # 统计错误类型分布
                error_type_summary = AnomalyService._format_error_types(stats['error_types'])
                anomaly = {
                    'type': 'universal_error',
                    'severity': 'critical',
                    'page_num': page_num,
                    'question_index': q_index,
                    'error_rate': error_rate,
                    'sample_count': stats['total'],
                    'correct_count': stats['correct'],
                    'error_count': stats['error'],
                    'error_types': dict(stats['error_types']),
                    'description': f'全部{stats["total"]}人判错',
                    'error_type_summary': error_type_summary,
                    'samples': stats['samples'][:5]
                }
                universal_errors += 1
            
            # 2. 有人对有人错的情况（不稳定）
            elif stats['correct'] > 0 and stats['error'] > 0 and stats['total'] >= 2:
                # 分析错误类型
                error_types = stats['error_types']
                high_anomaly_count = sum(error_types.get(t, 0) for t in AnomalyService.HIGH_ANOMALY_ERROR_TYPES)
                low_anomaly_count = stats['error'] - high_anomaly_count
                
                # 判断主要异常类型
                if high_anomaly_count > 0:
                    # 高异常：存在"识别正确-判断错误"
                    error_type_summary = AnomalyService._format_error_types(stats['error_types'])
                    anomaly = {
                        'type': 'high_anomaly',
                        'severity': 'high',
                        'page_num': page_num,
                        'question_index': q_index,
                        'error_rate': error_rate,
                        'sample_count': stats['total'],
                        'correct_count': stats['correct'],
                        'error_count': stats['error'],
                        'error_types': dict(stats['error_types']),
                        'description': f'{stats["correct"]}对/{stats["error"]}错，含识别正确-判断错误',
                        'error_type_summary': error_type_summary,
                        'samples': stats['samples'][:5]
                    }
                    high_anomaly += 1
                else:
                    # 低异常：其他错误类型
                    error_type_summary = AnomalyService._format_error_types(stats['error_types'])
                    anomaly = {
                        'type': 'low_anomaly',
                        'severity': 'medium',
                        'page_num': page_num,
                        'question_index': q_index,
                        'error_rate': error_rate,
                        'sample_count': stats['total'],
                        'correct_count': stats['correct'],
                        'error_count': stats['error'],
                        'error_types': dict(stats['error_types']),
                        'description': f'{stats["correct"]}对/{stats["error"]}错',
                        'error_type_summary': error_type_summary,
                        'samples': stats['samples'][:5]
                    }
                    low_anomaly += 1
            
            else:
                normal += 1
            
            if anomaly:
                anomalies.append(anomaly)
        
        # 按严重程度和错误率排序
        severity_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        anomalies.sort(key=lambda x: (severity_order.get(x['severity'], 99), -x['error_rate']))
        
        total_questions = len(question_stats)
        anomaly_count = universal_errors + high_anomaly + low_anomaly
        
        # 计算一致性（排除全员错误）
        effective_questions = total_questions - universal_errors  # 有效题数
        consistency_rate = round(normal / effective_questions, 4) if effective_questions > 0 else 0
        
        return {
            'summary': {
                'universal_errors': universal_errors,
                'high_anomaly': high_anomaly,
                'low_anomaly': low_anomaly,
                'normal': normal,
                'total_questions': total_questions,
                'anomaly_count': anomaly_count,
                'anomaly_rate': round(anomaly_count / total_questions, 4) if total_questions > 0 else 0,
                # 一致性相关
                'effective_questions': effective_questions,
                'consistency_rate': consistency_rate
            },
            'anomalies': anomalies
        }
    
    @staticmethod
    def _format_error_types(error_types: Dict[str, int]) -> str:
        """格式化错误类型统计为可读字符串"""
        if not error_types:
            return ''
        parts = [f'{t}({c})' for t, c in sorted(error_types.items(), key=lambda x: -x[1])]
        return '、'.join(parts[:3])  # 最多显示3种
    
    @staticmethod
    def _get_inconsistent_details(stats: Dict) -> Dict:
        """获取不稳定判断的详细信息（保留兼容）"""
        return {'description': '', 'inconsistent_answers': []}
    
    @staticmethod
    def detect_task_anomaly(task_id: str) -> Optional[Dict[str, Any]]:
        """
        检测任务准确率异常 (US-26.1)
        
        计算历史准确率的均值和标准差，
        判断当前任务准确率是否偏离均值超过阈值。
        
        Args:
            task_id: 批量任务ID
            
        Returns:
            dict: 异常信息，无异常返回 None
        """
        # 加载当前任务
        task_file = os.path.join(StorageService.BATCH_TASKS_DIR, f'{task_id}.json')
        if not os.path.exists(task_file):
            return None
        
        try:
            with open(task_file, 'r', encoding='utf-8') as f:
                task_data = json.load(f)
        except:
            return None
        
        overall_report = task_data.get('overall_report') or {}
        current_accuracy = overall_report.get('overall_accuracy', 0)
        
        # 获取历史准确率
        history_accuracies = AnomalyService._get_history_accuracies(task_id)
        
        if len(history_accuracies) < AnomalyService._config['min_samples']:
            return None  # 样本不足，无法检测
        
        # 计算统计值
        mean_acc = statistics.mean(history_accuracies)
        std_acc = statistics.stdev(history_accuracies)
        
        if std_acc == 0:
            return None  # 标准差为0，无法检测
        
        # 计算偏差
        threshold = AnomalyService._config['threshold_sigma']
        deviation = abs(current_accuracy - mean_acc) / std_acc
        
        if deviation <= threshold:
            return None  # 在正常范围内

        # 检测到异常，记录日志
        anomaly_type = 'accuracy_drop' if current_accuracy < mean_acc else 'accuracy_spike'
        severity = 'high' if deviation > threshold * 1.5 else 'medium'
        
        anomaly = {
            'anomaly_id': str(uuid.uuid4())[:8],
            'anomaly_type': anomaly_type,
            'severity': severity,
            'task_id': task_id,
            'metric_name': 'accuracy',
            'expected_value': round(mean_acc, 4),
            'actual_value': round(current_accuracy, 4),
            'deviation': round(deviation, 4),
            'threshold': threshold,
            'message': f'准确率异常: 期望 {mean_acc:.2%}，实际 {current_accuracy:.2%}，偏离 {deviation:.1f}σ'
        }
        
        # 保存到数据库
        AnomalyService._save_anomaly(anomaly)
        
        return anomaly
    
    @staticmethod
    def _get_history_accuracies(exclude_task_id: str = None, days: int = 30) -> List[float]:
        """获取历史准确率列表"""
        accuracies = []
        batch_tasks_dir = StorageService.BATCH_TASKS_DIR
        
        if not os.path.exists(batch_tasks_dir):
            return accuracies
        
        cutoff_date = datetime.now() - timedelta(days=days)
        
        for filename in os.listdir(batch_tasks_dir):
            if not filename.endswith('.json'):
                continue
            
            task_id = filename.replace('.json', '')
            if task_id == exclude_task_id:
                continue
            
            filepath = os.path.join(batch_tasks_dir, filename)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    task_data = json.load(f)
                
                # 检查时间
                created_at = task_data.get('created_at', '')
                if created_at:
                    try:
                        task_time = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        if task_time.tzinfo:
                            task_time = task_time.replace(tzinfo=None)
                        if task_time < cutoff_date:
                            continue
                    except:
                        pass
                
                # 获取准确率
                overall_report = task_data.get('overall_report') or {}
                accuracy = overall_report.get('overall_accuracy')
                if accuracy is not None:
                    accuracies.append(accuracy)
            except:
                continue
        
        return accuracies
    
    @staticmethod
    def _save_anomaly(anomaly: Dict[str, Any]) -> None:
        """保存异常到数据库"""
        sql = """
            INSERT INTO anomaly_logs 
            (anomaly_id, anomaly_type, severity, task_id, metric_name,
             expected_value, actual_value, deviation, threshold, message)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        AppDatabaseService.execute_insert(sql, (
            anomaly['anomaly_id'], anomaly['anomaly_type'], anomaly['severity'],
            anomaly.get('task_id'), anomaly.get('metric_name'),
            anomaly.get('expected_value'), anomaly.get('actual_value'),
            anomaly.get('deviation'), anomaly.get('threshold'),
            anomaly.get('message')
        ))

    
    @staticmethod
    def get_anomaly_logs(
        page: int = 1,
        page_size: int = 20,
        anomaly_type: str = None,
        severity: str = None,
        is_acknowledged: bool = None
    ) -> Dict[str, Any]:
        """
        获取异常日志列表 (US-26.5)
        
        Args:
            page: 页码
            page_size: 每页数量
            anomaly_type: 异常类型筛选
            severity: 严重程度筛选
            is_acknowledged: 是否已确认筛选
            
        Returns:
            dict: {items: list, total: int, page: int, page_size: int}
        """
        where_clauses = ['1=1']
        params = []
        
        if anomaly_type:
            where_clauses.append('anomaly_type = %s')
            params.append(anomaly_type)
        
        if severity:
            where_clauses.append('severity = %s')
            params.append(severity)
        
        if is_acknowledged is not None:
            where_clauses.append('is_acknowledged = %s')
            params.append(1 if is_acknowledged else 0)
        
        where_sql = ' AND '.join(where_clauses)
        
        # 查询总数
        count_sql = f'SELECT COUNT(*) as total FROM anomaly_logs WHERE {where_sql}'
        count_result = AppDatabaseService.execute_one(count_sql, tuple(params) if params else None)
        total = count_result['total'] if count_result else 0
        
        # 查询列表
        offset = (page - 1) * page_size
        list_sql = f"""
            SELECT anomaly_id, anomaly_type, severity, task_id, metric_name,
                   expected_value, actual_value, deviation, threshold, message,
                   is_acknowledged, acknowledged_by, acknowledged_at, created_at
            FROM anomaly_logs 
            WHERE {where_sql}
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        """
        params.extend([page_size, offset])
        rows = AppDatabaseService.execute_query(list_sql, tuple(params))
        
        items = []
        for row in rows:
            items.append({
                'anomaly_id': row['anomaly_id'],
                'anomaly_type': row['anomaly_type'],
                'severity': row['severity'],
                'task_id': row.get('task_id'),
                'metric_name': row.get('metric_name'),
                'expected_value': float(row['expected_value']) if row.get('expected_value') else None,
                'actual_value': float(row['actual_value']) if row.get('actual_value') else None,
                'deviation': float(row['deviation']) if row.get('deviation') else None,
                'threshold': float(row['threshold']) if row.get('threshold') else None,
                'message': row.get('message', ''),
                'is_acknowledged': bool(row.get('is_acknowledged')),
                'acknowledged_by': row.get('acknowledged_by'),
                'acknowledged_at': row['acknowledged_at'].isoformat() if row.get('acknowledged_at') else None,
                'created_at': row['created_at'].isoformat() if row.get('created_at') else ''
            })
        
        return {
            'items': items,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    
    @staticmethod
    def acknowledge_anomaly(anomaly_id: str, user_id: int = None) -> bool:
        """确认异常"""
        sql = """
            UPDATE anomaly_logs 
            SET is_acknowledged = 1, acknowledged_by = %s, acknowledged_at = NOW()
            WHERE anomaly_id = %s
        """
        result = AppDatabaseService.execute_update(sql, (user_id, anomaly_id))
        return result > 0
    
    @staticmethod
    def get_statistics() -> Dict[str, Any]:
        """获取异常统计"""
        # 未确认数量
        unack_sql = "SELECT COUNT(*) as count FROM anomaly_logs WHERE is_acknowledged = 0"
        unack_result = AppDatabaseService.execute_one(unack_sql)
        unacknowledged = unack_result['count'] if unack_result else 0
        
        # 今日异常数
        today_sql = """
            SELECT COUNT(*) as count FROM anomaly_logs 
            WHERE DATE(created_at) = CURDATE()
        """
        today_result = AppDatabaseService.execute_one(today_sql)
        today_count = today_result['count'] if today_result else 0
        
        # 按类型统计
        type_sql = """
            SELECT anomaly_type, COUNT(*) as count 
            FROM anomaly_logs 
            GROUP BY anomaly_type
        """
        type_rows = AppDatabaseService.execute_query(type_sql)
        by_type = {row['anomaly_type']: row['count'] for row in type_rows}
        
        # 按严重程度统计
        severity_sql = """
            SELECT severity, COUNT(*) as count 
            FROM anomaly_logs 
            GROUP BY severity
        """
        severity_rows = AppDatabaseService.execute_query(severity_sql)
        by_severity = {row['severity']: row['count'] for row in severity_rows}
        
        return {
            'unacknowledged': unacknowledged,
            'today_count': today_count,
            'by_type': by_type,
            'by_severity': by_severity
        }
    
    @staticmethod
    def set_threshold(threshold_sigma: float) -> None:
        """设置异常阈值 (US-26.3)"""
        if threshold_sigma < 1 or threshold_sigma > 5:
            raise ValueError('阈值必须在 1-5 之间')
        AnomalyService._config['threshold_sigma'] = threshold_sigma
    
    @staticmethod
    def _make_bar(value: int, max_value: int, width: int = 20) -> str:
        """生成文本进度条"""
        if max_value <= 0:
            return ''
        filled = int(value / max_value * width)
        return '█' * filled + '░' * (width - filled)
    
    @staticmethod
    def export_question_anomalies_to_excel(task_id: str) -> Optional[str]:
        """
        导出任务题目异常检测结果到Excel（纯黑白简洁风格）
        
        Args:
            task_id: 任务ID
            
        Returns:
            导出文件路径，失败返回 None
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        # 加载任务数据
        task_data = StorageService.load_batch_task(task_id)
        if not task_data:
            return None
        
        # 检测异常
        result = AnomalyService.detect_question_anomalies(task_data)
        summary = result.get('summary', {})
        anomalies = result.get('anomalies', [])
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "汇总"
        
        # 样式定义（纯黑白）
        font_title = Font(bold=True, size=14)
        font_subtitle = Font(bold=True, size=11)
        font_header = Font(bold=True, size=10)
        font_normal = Font(size=10)
        font_small = Font(size=9)
        font_big = Font(bold=True, size=18)
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        border_bottom = Border(bottom=Side(style='thin', color='000000'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')
        
        # 异常类型映射
        type_map = {
            'universal_error': '全员错误',
            'high_anomaly': '高异常',
            'low_anomaly': '低异常'
        }
        severity_map = {
            'critical': '严重',
            'high': '高',
            'medium': '中',
            'low': '低'
        }
        
        # 提取数据
        total = summary.get('total_questions', 0)
        normal = summary.get('normal', 0)
        universal = summary.get('universal_errors', 0)
        high_anom = summary.get('high_anomaly', 0)
        low_anom = summary.get('low_anomaly', 0)
        effective = summary.get('effective_questions', total - universal)
        consistency = summary.get('consistency_rate', 0)
        anomaly_total = universal + high_anom + low_anom
        
        # 统计作业数
        homework_items = task_data.get('homework_items', [])
        completed_count = len([h for h in homework_items if h.get('status') == 'completed'])
        
        # 按页码统计异常
        page_anomaly_stats = {}
        for a in anomalies:
            page = str(a.get('page_num', '?'))
            if page not in page_anomaly_stats:
                page_anomaly_stats[page] = {'universal': 0, 'high': 0, 'low': 0, 'total': 0, 'items': []}
            page_anomaly_stats[page]['total'] += 1
            atype = a.get('type', '')
            if atype == 'universal_error':
                page_anomaly_stats[page]['universal'] += 1
            elif atype == 'high_anomaly':
                page_anomaly_stats[page]['high'] += 1
            else:
                page_anomaly_stats[page]['low'] += 1
            page_anomaly_stats[page]['items'].append(a)
        
        # ========== 汇总工作表（仪表盘布局）==========
        # 设置列宽
        col_widths = [12, 18, 8, 10, 12, 6, 22, 8, 10, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        row = 1
        
        # === 标题行 ===
        task_name = task_data.get('name', task_id[:8])
        ws.cell(row=row, column=1, value="异常检测分析报告").font = font_title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        
        # === 基本信息行 ===
        info_text = f"任务: {task_name}    检测时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}    作业数: {completed_count}份"
        ws.cell(row=row, column=1, value=info_text).font = font_small
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2
        
        # === 左侧：整体评估 | 右侧：异常分布 ===
        ws.cell(row=row, column=1, value="整体评估").font = font_subtitle
        ws.cell(row=row, column=6, value="异常分布").font = font_subtitle
        row += 1
        
        # 一致性进度条
        bar = AnomalyService._make_bar(int(consistency * 100), 100, 20)
        ws.cell(row=row, column=1, value=f"{bar} {consistency*100:.1f}%").font = font_normal
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        
        # 异常分布 - 全员错误
        ws.cell(row=row, column=6, value="全员错误").font = font_normal
        ws.cell(row=row, column=7, value=AnomalyService._make_bar(universal, max(anomaly_total, 1), 15)).font = font_normal
        ws.cell(row=row, column=8, value=f"{universal}题").font = font_normal
        ws.cell(row=row, column=9, value="! 检查标注").font = font_small
        row += 1
        
        ws.cell(row=row, column=1, value=f"一致性 ({normal}/{effective}有效题)").font = font_small
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        
        # 异常分布 - 高异常
        ws.cell(row=row, column=6, value="高异常").font = font_normal
        ws.cell(row=row, column=7, value=AnomalyService._make_bar(high_anom, max(anomaly_total, 1), 15)).font = font_normal
        ws.cell(row=row, column=8, value=f"{high_anom}题").font = font_normal
        ws.cell(row=row, column=9, value="! 判断逻辑").font = font_small
        row += 1
        
        # 题目统计
        ws.cell(row=row, column=1, value=f"总题数: {total}  异常: {anomaly_total}  正常: {normal}").font = font_normal
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        
        # 异常分布 - 低异常
        ws.cell(row=row, column=6, value="低异常").font = font_normal
        ws.cell(row=row, column=7, value=AnomalyService._make_bar(low_anom, max(anomaly_total, 1), 15)).font = font_normal
        ws.cell(row=row, column=8, value=f"{low_anom}题").font = font_normal
        ws.cell(row=row, column=9, value="o 可接受").font = font_small
        row += 2
        
        # === 左侧：关键发现 | 右侧：页码热点 ===
        ws.cell(row=row, column=1, value="关键发现").font = font_subtitle
        ws.cell(row=row, column=6, value="页码热点 TOP3").font = font_subtitle
        row += 1
        
        # 关键发现
        findings = []
        if universal > 0:
            # 找出全员错误的题目
            ue_items = [a for a in anomalies if a.get('type') == 'universal_error'][:3]
            ue_refs = ', '.join([f"P{a.get('page_num')}#{a.get('question_index')}" for a in ue_items])
            findings.append(f"! {universal}道全员错误 -> {ue_refs}")
        if high_anom > 0:
            findings.append(f"! {high_anom}道高异常(识别对判断错)")
        if consistency >= 0.9:
            findings.append(f"v 一致性{consistency*100:.1f}%，表现良好")
        elif consistency < 0.7:
            findings.append(f"! 一致性仅{consistency*100:.1f}%，需排查")
        else:
            findings.append(f"o 一致性{consistency*100:.1f}%，基本正常")
        if not findings:
            findings.append("v 未发现明显异常")
        
        # 页码热点排序
        sorted_pages = sorted(page_anomaly_stats.items(), key=lambda x: -x[1]['total'])[:3]
        max_page_anomaly = sorted_pages[0][1]['total'] if sorted_pages else 1
        
        for i, finding in enumerate(findings[:3]):
            ws.cell(row=row + i, column=1, value=finding).font = font_normal
            ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=4)
        
        for i, (page, stats) in enumerate(sorted_pages):
            sev = "[严重]" if stats['universal'] > 0 else ("[较高]" if stats['high'] > 0 else "[一般]")
            ws.cell(row=row + i, column=6, value=f"P{page}").font = font_normal
            ws.cell(row=row + i, column=7, value=AnomalyService._make_bar(stats['total'], max_page_anomaly, 10)).font = font_normal
            ws.cell(row=row + i, column=8, value=f"{stats['total']}题").font = font_normal
            ws.cell(row=row + i, column=9, value=sev).font = font_small
        
        row += max(len(findings), len(sorted_pages), 3) + 1
        
        # === 问题题目速查表 ===
        ws.cell(row=row, column=1, value="问题题目速查 (严重/高异常)").font = font_subtitle
        row += 1
        
        # 表头
        quick_headers = ['严重度', '页码', '题号', '问题类型', '错误人数', '说明']
        for col, h in enumerate(quick_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = font_header
            cell.border = border_thin
            cell.alignment = align_center
        row += 1
        
        # 只显示严重和高异常
        critical_anomalies = [a for a in anomalies if a.get('severity') in ('critical', 'high')][:10]
        for a in critical_anomalies:
            correct = a.get('correct_count', 0)
            error = a.get('error_count', 0)
            sample_count = a.get('sample_count', correct + error)
            
            ws.cell(row=row, column=1, value=severity_map.get(a.get('severity', ''), '')).border = border_thin
            ws.cell(row=row, column=2, value=f"P{a.get('page_num', '')}").border = border_thin
            ws.cell(row=row, column=3, value=str(a.get('question_index', ''))).border = border_thin
            ws.cell(row=row, column=4, value=type_map.get(a.get('type', ''), '')).border = border_thin
            ws.cell(row=row, column=5, value=f"{error}/{sample_count}").border = border_thin
            
            # 说明
            if a.get('type') == 'universal_error':
                desc = "所有人判错，检查标注"
            elif a.get('type') == 'high_anomaly':
                desc = "识别正确-判断错误"
            else:
                desc = a.get('error_type_summary', '')[:20]
            ws.cell(row=row, column=6, value=desc).border = border_thin
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).font = font_normal
                ws.cell(row=row, column=col).alignment = align_center if col <= 5 else align_left
            row += 1
        
        if not critical_anomalies:
            ws.cell(row=row, column=1, value="无严重/高异常题目").font = font_small
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
        
        # ========== 异常详情工作表 ==========
        ws_detail = wb.create_sheet("异常详情")
        
        # 设置列宽
        detail_col_widths = [8, 10, 12, 10, 12, 12, 35, 45]
        for i, width in enumerate(detail_col_widths, 1):
            ws_detail.column_dimensions[get_column_letter(i)].width = width
        
        # 表头
        headers = ['页码', '题号', '异常类型', '严重度', '正确/错误', '错误率', '错误类型', '样本详情']
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = font_header
            cell.border = border_thin
            cell.alignment = align_center
        
        # 数据行
        for row_idx, anomaly in enumerate(anomalies, 2):
            correct = anomaly.get('correct_count', 0)
            error = anomaly.get('error_count', 0)
            
            ws_detail.cell(row=row_idx, column=1, value=anomaly.get('page_num', ''))
            ws_detail.cell(row=row_idx, column=2, value=anomaly.get('question_index', ''))
            ws_detail.cell(row=row_idx, column=3, value=type_map.get(anomaly.get('type', ''), anomaly.get('type', '')))
            ws_detail.cell(row=row_idx, column=4, value=severity_map.get(anomaly.get('severity', ''), anomaly.get('severity', '')))
            ws_detail.cell(row=row_idx, column=5, value=f"{correct}对/{error}错")
            ws_detail.cell(row=row_idx, column=6, value=f"{anomaly.get('error_rate', 0) * 100:.0f}%")
            ws_detail.cell(row=row_idx, column=7, value=anomaly.get('error_type_summary', ''))
            
            # 样本详情
            samples = anomaly.get('samples', [])
            sample_texts = []
            for s in samples[:3]:
                student = s.get('student_name', s.get('student_id', ''))[:6]
                result_text = '对' if s.get('result') == 'correct' else '错'
                err_type = s.get('error_type', '')[:8]
                sample_texts.append(f"{student}:{result_text}" + (f"({err_type})" if err_type else ''))
            ws_detail.cell(row=row_idx, column=8, value='; '.join(sample_texts))
            
            for col in range(1, 9):
                cell = ws_detail.cell(row=row_idx, column=col)
                cell.font = font_normal
                cell.border = border_thin
                cell.alignment = align_center if col <= 6 else align_left
        
        # 冻结首行
        ws_detail.freeze_panes = 'A2'
        if anomalies:
            ws_detail.auto_filter.ref = f"A1:H{len(anomalies) + 1}"
        
        # ========== 按页码汇总工作表 ==========
        ws_page = wb.create_sheet("按页码汇总")
        
        # 设置列宽
        for i, w in enumerate([10, 12, 12, 12, 12, 18], 1):
            ws_page.column_dimensions[get_column_letter(i)].width = w
        
        # 表头
        page_headers = ['页码', '全员错误', '高异常', '低异常', '异常总数', '问题严重度']
        for col, h in enumerate(page_headers, 1):
            cell = ws_page.cell(row=1, column=col, value=h)
            cell.font = font_header
            cell.border = border_thin
            cell.alignment = align_center
        
        # 数据行
        sorted_page_list = sorted(page_anomaly_stats.keys(), key=lambda x: int(x) if x.isdigit() else 999)
        for row_idx, page in enumerate(sorted_page_list, 2):
            stats = page_anomaly_stats[page]
            ws_page.cell(row=row_idx, column=1, value=f"P{page}")
            ws_page.cell(row=row_idx, column=2, value=stats['universal'])
            ws_page.cell(row=row_idx, column=3, value=stats['high'])
            ws_page.cell(row=row_idx, column=4, value=stats['low'])
            ws_page.cell(row=row_idx, column=5, value=stats['total'])
            
            # 严重度评估（纯文本）
            if stats['universal'] > 0:
                sev_text = '[严重] 有全员错误'
            elif stats['high'] > 0:
                sev_text = '[较高] 有高异常'
            elif stats['low'] > 0:
                sev_text = '[一般] 仅低异常'
            else:
                sev_text = '[正常]'
            
            ws_page.cell(row=row_idx, column=6, value=sev_text)
            
            for col in range(1, 7):
                cell = ws_page.cell(row=row_idx, column=col)
                cell.font = font_normal
                cell.border = border_thin
                cell.alignment = align_center
        
        ws_page.freeze_panes = 'A2'
        
        # ========== 样本明细工作表 ==========
        ws_samples = wb.create_sheet("样本明细")
        
        # 收集所有样本
        all_samples = []
        for anomaly in anomalies:
            page_num = anomaly.get('page_num', '')
            q_index = anomaly.get('question_index', '')
            anomaly_type = type_map.get(anomaly.get('type', ''), anomaly.get('type', ''))
            for sample in anomaly.get('samples', []):
                all_samples.append({
                    'page_num': page_num,
                    'question_index': q_index,
                    'anomaly_type': anomaly_type,
                    'student_id': sample.get('student_id', ''),
                    'student_name': sample.get('student_name', ''),
                    'result': '正确' if sample.get('result') == 'correct' else '错误',
                    'hw_answer': sample.get('hw_answer', ''),
                    'base_answer': sample.get('base_answer', ''),
                    'base_user': sample.get('base_user', ''),
                    'error_type': sample.get('error_type', '')
                })
        
        # 设置列宽
        sample_col_widths = [8, 8, 10, 12, 12, 8, 20, 20, 20, 15]
        for i, width in enumerate(sample_col_widths, 1):
            ws_samples.column_dimensions[get_column_letter(i)].width = width
        
        # 表头
        sample_headers = ['页码', '题号', '异常类型', '学生ID', '学生姓名', '结果', 'AI识别答案', '标准答案', '基准用户答案', '错误类型']
        for col, header in enumerate(sample_headers, 1):
            cell = ws_samples.cell(row=1, column=col, value=header)
            cell.font = font_header
            cell.border = border_thin
            cell.alignment = align_center
        
        # 数据行
        for row_idx, sample in enumerate(all_samples, 2):
            ws_samples.cell(row=row_idx, column=1, value=sample.get('page_num', ''))
            ws_samples.cell(row=row_idx, column=2, value=sample.get('question_index', ''))
            ws_samples.cell(row=row_idx, column=3, value=sample.get('anomaly_type', ''))
            ws_samples.cell(row=row_idx, column=4, value=sample.get('student_id', ''))
            ws_samples.cell(row=row_idx, column=5, value=sample.get('student_name', ''))
            ws_samples.cell(row=row_idx, column=6, value=sample.get('result', ''))
            ws_samples.cell(row=row_idx, column=7, value=str(sample.get('hw_answer', ''))[:50])
            ws_samples.cell(row=row_idx, column=8, value=str(sample.get('base_answer', ''))[:50])
            ws_samples.cell(row=row_idx, column=9, value=str(sample.get('base_user', ''))[:50])
            ws_samples.cell(row=row_idx, column=10, value=sample.get('error_type', ''))
            
            for col in range(1, 11):
                cell = ws_samples.cell(row=row_idx, column=col)
                cell.font = font_normal
                cell.border = border_thin
                cell.alignment = align_center if col <= 6 else align_left
        
        # 冻结首行
        ws_samples.freeze_panes = 'A2'
        if all_samples:
            ws_samples.auto_filter.ref = f"A1:J{len(all_samples) + 1}"
        
        # 保存文件
        task_name_safe = task_name[:20].replace('/', '_').replace('\\', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"异常检测报告_{task_name_safe}_{timestamp}.xlsx"
        filepath = os.path.join('exports', filename)
        
        os.makedirs('exports', exist_ok=True)
        wb.save(filepath)
        
        return filepath
