"""
Excel导出服务模块
提供批量评估Excel报告的数据提取、样式配置和工作表生成功能
"""
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, BarChart, LineChart, RadarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from services.storage_service import StorageService


# ========== 安全数据获取工具函数 ==========

def safe_get(data: dict, key: str, default: Any = '') -> Any:
    """安全获取字典值，缺失时返回默认值"""
    if not data or not isinstance(data, dict):
        return default
    return data.get(key, default) if data.get(key) is not None else default


def safe_get_nested(data: dict, keys: List[str], default: Any = '') -> Any:
    """安全获取嵌套字典值"""
    try:
        result = data
        for key in keys:
            if isinstance(result, dict):
                result = result.get(key)
            elif isinstance(result, list) and isinstance(key, int):
                result = result[key] if 0 <= key < len(result) else None
            else:
                return default
            if result is None:
                return default
        return result if result is not None else default
    except (KeyError, TypeError, IndexError):
        return default


def get_with_fallback(item: dict, keys: List[str], default: Any = '') -> Any:
    """尝试多个键获取值，支持向后兼容"""
    if not item or not isinstance(item, dict):
        return default
    for key in keys:
        if key in item and item[key] is not None and item[key] != '':
            return item[key]
    return default


def safe_float(value: Any, default: float = 0.0) -> float:
    """安全转换为浮点数"""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_int(value: Any, default: int = 0) -> int:
    """安全转换为整数"""
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def safe_json_loads(json_str: Any, default: Any = None) -> Any:
    """安全解析JSON"""
    if default is None:
        default = []
    if not json_str:
        return default
    if isinstance(json_str, (list, dict)):
        return json_str
    try:
        return json.loads(json_str)
    except (json.JSONDecodeError, TypeError):
        return default


def normalize_homework_result(homework_result: Any) -> List[Dict]:
    """标准化作业结果数据结构"""
    result = safe_json_loads(homework_result, [])
    if isinstance(result, dict):
        return [result]
    if not isinstance(result, list):
        return []
    return result


def get_correct_value(item: dict) -> str:
    """获取判断结果字段值，统一返回 'yes' 或 'no'"""
    if not item:
        return ''
    
    for field in ['correct', 'isRight', 'isCorrect']:
        if field in item:
            val = item[field]
            if isinstance(val, bool):
                return 'yes' if val else 'no'
            if isinstance(val, str):
                val_lower = val.lower()
                return 'yes' if val_lower in ('yes', 'true', '1') else 'no'
    return ''


def classify_question_type(question_data: dict) -> Dict[str, Any]:
    """根据题目数据判断题目类型（三类互不包含）"""
    if not question_data:
        return {'is_choice': False, 'is_fill': False, 'is_subjective': True, 'is_parent': False, 'type_name': '主观题'}
    
    bvalue = str(question_data.get('bvalue', ''))
    question_type = question_data.get('questionType', '')
    children = question_data.get('children', [])
    
    if children:
        return {'is_choice': False, 'is_fill': False, 'is_subjective': False, 'is_parent': True, 'type_name': '大题'}
    
    is_choice = bvalue in ('1', '2', '3')
    is_fill = (question_type == 'objective' and bvalue == '4')
    is_subjective = not is_choice and not is_fill
    
    type_name = '选择题' if is_choice else ('客观填空题' if is_fill else '主观题')
    return {'is_choice': is_choice, 'is_fill': is_fill, 'is_subjective': is_subjective, 'is_parent': False, 'type_name': type_name}


# ========== 样式配置类 ==========

class StyleConfig:
    """Excel样式配置 - 黑白简洁风格"""
    
    # 颜色定义
    COLOR_BLACK = "1D1D1F"
    COLOR_WHITE = "FFFFFF"
    COLOR_SUCCESS = "E3F9E5"
    COLOR_WARNING = "FFF3E0"
    COLOR_ERROR = "FFEEF0"
    COLOR_INFO = "E3F2FD"
    
    # 字体
    FONT_TITLE = Font(bold=True, size=16, color="1D1D1F")
    FONT_SUBTITLE = Font(bold=True, size=14, color="1D1D1F")
    FONT_HEADER = Font(bold=True, size=12, color="FFFFFF")
    FONT_SECTION = Font(bold=True, size=12, color="1D1D1F")
    FONT_NORMAL = Font(size=11, color="1D1D1F")
    
    # 填充
    FILL_HEADER = PatternFill(start_color="1D1D1F", end_color="1D1D1F", fill_type="solid")
    FILL_SUCCESS = PatternFill(start_color="E3F9E5", end_color="E3F9E5", fill_type="solid")
    FILL_WARNING = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    FILL_ERROR = PatternFill(start_color="FFEEF0", end_color="FFEEF0", fill_type="solid")
    FILL_INFO = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # 对齐
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 边框
    BORDER_THIN = Border(
        left=Side(style='thin', color='D2D2D7'),
        right=Side(style='thin', color='D2D2D7'),
        top=Side(style='thin', color='D2D2D7'),
        bottom=Side(style='thin', color='D2D2D7')
    )
    
    @classmethod
    def get_accuracy_fill(cls, accuracy: float) -> PatternFill:
        """根据准确率返回对应的填充颜色"""
        if accuracy >= 0.9:
            return cls.FILL_SUCCESS
        elif accuracy >= 0.7:
            return cls.FILL_WARNING
        return cls.FILL_ERROR
    
    @classmethod
    def apply_header_style(cls, cell):
        """应用表头样式"""
        cell.font = cls.FONT_HEADER
        cell.fill = cls.FILL_HEADER
        cell.border = cls.BORDER_THIN
        cell.alignment = cls.ALIGN_CENTER
    
    @classmethod
    def apply_cell_style(cls, cell, center: bool = False):
        """应用普通单元格样式"""
        cell.font = cls.FONT_NORMAL
        cell.border = cls.BORDER_THIN
        cell.alignment = cls.ALIGN_CENTER if center else cls.ALIGN_LEFT


# ========== 数据提取类 ==========

class DataExtractor:
    """数据提取器 - 从任务数据中提取各种统计信息"""
    
    @staticmethod
    def extract_error_details(homework_items: List[Dict], task_data: Dict) -> List[Dict]:
        """提取完整的错误详情数据"""
        error_details = []
        
        for item in homework_items:
            if item.get('status') != 'completed':
                continue
            
            evaluation = item.get('evaluation', {})
            errors = evaluation.get('errors', [])
            
            # 获取基准效果和AI结果映射
            base_effects = DataExtractor._get_base_effects_for_homework(item, task_data)
            base_map = {str(q.get('index', '')): q for q in base_effects}
            
            homework_result = normalize_homework_result(item.get('homework_result', '[]'))
            ai_map = {}
            for q in homework_result:
                ai_map[str(q.get('index', ''))] = q
                for child in q.get('children', []):
                    ai_map[str(child.get('index', ''))] = child
            
            for err in errors:
                idx = str(err.get('index', ''))
                base_item = base_map.get(idx, {})
                ai_item = ai_map.get(idx, err.get('ai_result', {}))
                
                error_details.append({
                    'homework_id': item.get('homework_id', ''),
                    'book_name': item.get('book_name', ''),
                    'page_num': item.get('page_num', ''),
                    'student_name': get_with_fallback(item, ['student_name', 'studentName'], '未知学生'),
                    'student_id': get_with_fallback(item, ['student_id', 'studentId'], ''),
                    'index': idx,
                    'error_type': err.get('error_type', '未分类'),
                    'base_user_answer': get_with_fallback(base_item, ['userAnswer', 'user_answer'], ''),
                    'ai_user_answer': get_with_fallback(ai_item, ['userAnswer', 'user_answer'], ''),
                    'base_correct': get_correct_value(base_item),
                    'ai_correct': get_correct_value(ai_item),
                    'standard_answer': get_with_fallback(base_item, ['answer', 'mainAnswer'], ''),
                    'similarity': err.get('similarity'),
                    'severity': err.get('severity', 'medium'),
                    'explanation': err.get('explanation', ''),
                    'question_type': classify_question_type(base_item).get('type_name', '未分类')
                })
        
        return error_details
    
    @staticmethod
    def _get_base_effects_for_homework(item: Dict, task_data: Dict) -> List[Dict]:
        """获取作业的基准效果数据"""
        base_effects = []
        matched_dataset = item.get('matched_dataset')
        
        if matched_dataset:
            ds_data = StorageService.load_dataset(matched_dataset)
            if ds_data:
                page_key = str(item.get('page_num', ''))
                base_effects = ds_data.get('base_effects', {}).get(page_key, [])
        
        if not base_effects:
            base_effects = task_data.get('base_effects', {}).get(str(item.get('page_num', '')), [])
        
        # 展开children
        flattened = []
        for q in base_effects:
            children = q.get('children', [])
            flattened.extend(children if children else [q])
        return flattened

    
    @staticmethod
    def extract_student_statistics(homework_items: List[Dict]) -> List[Dict]:
        """按学生统计数据"""
        student_stats = {}
        
        for item in homework_items:
            if item.get('status') != 'completed':
                continue
            
            student_name = get_with_fallback(item, ['student_name', 'studentName'], '未知学生')
            student_id = get_with_fallback(item, ['student_id', 'studentId'], '')
            key = f"{student_id}_{student_name}" if student_id else student_name
            
            if key not in student_stats:
                student_stats[key] = {
                    'student_name': student_name, 'student_id': student_id,
                    'homework_count': 0, 'total_questions': 0, 'correct_count': 0, 'error_count': 0,
                    'choice_total': 0, 'choice_correct': 0,
                    'fill_total': 0, 'fill_correct': 0,
                    'subjective_total': 0, 'subjective_correct': 0
                }
            
            stats = student_stats[key]
            stats['homework_count'] += 1
            
            evaluation = item.get('evaluation', {})
            stats['total_questions'] += evaluation.get('total_questions', 0)
            stats['correct_count'] += evaluation.get('correct_count', 0)
            stats['error_count'] += evaluation.get('error_count', 0)
            
            by_type = evaluation.get('by_question_type', {})
            for type_key, stat_key in [('choice', 'choice'), ('objective_fill', 'fill'), ('subjective', 'subjective')]:
                type_data = by_type.get(type_key, by_type.get('other', {})) if type_key == 'subjective' else by_type.get(type_key, {})
                stats[f'{stat_key}_total'] += type_data.get('total', 0)
                stats[f'{stat_key}_correct'] += type_data.get('correct', 0)
        
        result = []
        for stats in student_stats.values():
            total = stats['total_questions']
            stats['accuracy'] = stats['correct_count'] / total if total > 0 else 0
            stats['choice_accuracy'] = stats['choice_correct'] / stats['choice_total'] if stats['choice_total'] > 0 else 0
            stats['fill_accuracy'] = stats['fill_correct'] / stats['fill_total'] if stats['fill_total'] > 0 else 0
            stats['subjective_accuracy'] = stats['subjective_correct'] / stats['subjective_total'] if stats['subjective_total'] > 0 else 0
            result.append(stats)
        
        result.sort(key=lambda x: x['accuracy'], reverse=True)
        return result
    
    @staticmethod
    def extract_page_statistics(homework_items: List[Dict]) -> List[Dict]:
        """按页码统计数据"""
        page_stats = {}
        
        for item in homework_items:
            if item.get('status') != 'completed':
                continue
            
            page_num = str(item.get('page_num', '?'))
            book_name = item.get('book_name', '')
            key = f"{book_name}_{page_num}"
            
            if key not in page_stats:
                page_stats[key] = {
                    'book_name': book_name, 'page_num': page_num,
                    'homework_count': 0, 'total_questions': 0, 'correct_count': 0, 'error_count': 0
                }
            
            stats = page_stats[key]
            stats['homework_count'] += 1
            evaluation = item.get('evaluation', {})
            stats['total_questions'] += evaluation.get('total_questions', 0)
            stats['correct_count'] += evaluation.get('correct_count', 0)
            stats['error_count'] += evaluation.get('error_count', 0)
        
        result = []
        for stats in page_stats.values():
            stats['accuracy'] = stats['correct_count'] / stats['total_questions'] if stats['total_questions'] > 0 else 0
            result.append(stats)
        
        result.sort(key=lambda x: safe_int(x['page_num'], 999))
        return result

    
    @staticmethod
    def extract_type_statistics(homework_items: List[Dict]) -> List[Dict]:
        """按题型统计数据"""
        type_stats = {
            'choice': {'type_name': '选择题', 'total': 0, 'correct': 0},
            'objective_fill': {'type_name': '客观填空题', 'total': 0, 'correct': 0},
            'subjective': {'type_name': '主观题', 'total': 0, 'correct': 0}
        }
        
        for item in homework_items:
            if item.get('status') != 'completed':
                continue
            
            by_type = item.get('evaluation', {}).get('by_question_type', {})
            for key in type_stats:
                data = by_type.get(key, by_type.get('other', {})) if key == 'subjective' else by_type.get(key, {})
                type_stats[key]['total'] += data.get('total', 0)
                type_stats[key]['correct'] += data.get('correct', 0)
        
        result = []
        for stats in type_stats.values():
            stats['error_count'] = stats['total'] - stats['correct']
            stats['accuracy'] = stats['correct'] / stats['total'] if stats['total'] > 0 else 0
            result.append(stats)
        return result
    
    @staticmethod
    def extract_essay_scores(homework_items: List[Dict], subject_id: int) -> Dict:
        """提取英语作文评分数据"""
        import re
        
        if subject_id != 0:
            return {'has_essay': False, 'essays': [], 'stats': None}
        
        essays, scores = [], []
        
        for item in homework_items:
            for q in normalize_homework_result(item.get('homework_result', '[]')):
                for target in [q] + q.get('children', []):
                    main_answer = target.get('mainAnswer', '')
                    if main_answer and '参考得分' in main_answer:
                        parsed = DataExtractor._parse_essay_feedback(main_answer)
                        if parsed and parsed.get('score') is not None:
                            essays.append({
                                'homework_id': item.get('homework_id', ''),
                                'student_name': get_with_fallback(item, ['student_name', 'studentName'], '未知学生'),
                                'student_id': get_with_fallback(item, ['student_id', 'studentId'], ''),
                                'index': target.get('index', ''),
                                'score': parsed['score'],
                                'evaluation': parsed['evaluation'],
                                'suggestions': parsed['suggestions']
                            })
                            scores.append(parsed['score'])
        
        if not essays:
            return {'has_essay': False, 'essays': [], 'stats': None}
        
        return {
            'has_essay': True, 'essays': essays,
            'stats': {
                'count': len(essays),
                'avg_score': round(sum(scores) / len(scores), 2),
                'max_score': max(scores),
                'min_score': min(scores)
            }
        }
    
    @staticmethod
    def _parse_essay_feedback(main_answer: str) -> Optional[Dict]:
        """解析作文评分反馈"""
        import re
        if not main_answer:
            return None
        
        result = {'score': None, 'evaluation': '', 'suggestions': ''}
        
        match = re.search(r'参考得分[：:]\s*([\d.]+)', main_answer)
        if match:
            try:
                result['score'] = float(match.group(1))
            except:
                pass
        
        eval_match = re.search(r'综合评价[：:]\s*(.+?)(?=针对性改进建议|改进建议|$)', main_answer, re.DOTALL)
        if eval_match:
            result['evaluation'] = eval_match.group(1).strip()[:300]
        
        sug_match = re.search(r'(?:针对性)?改进建议[：:]\s*(.+)', main_answer, re.DOTALL)
        if sug_match:
            result['suggestions'] = sug_match.group(1).strip()[:300]
        
        return result

    
    @staticmethod
    def extract_dataset_info(task_data: Dict) -> Optional[Dict]:
        """提取数据集信息"""
        homework_items = task_data.get('homework_items', [])
        dataset_ids = set(item.get('matched_dataset') for item in homework_items if item.get('matched_dataset'))
        
        if not dataset_ids:
            return None
        
        datasets = []
        for ds_id in dataset_ids:
            ds_data = StorageService.load_dataset(ds_id)
            if ds_data:
                base_effects = ds_data.get('base_effects', {})
                datasets.append({
                    'dataset_id': ds_id,
                    'name': ds_data.get('name', ds_id[:8]),
                    'description': ds_data.get('description', ''),
                    'created_at': ds_data.get('created_at', ''),
                    'pages': list(base_effects.keys()),
                    'question_count': sum(len(q) for q in base_effects.values())
                })
        
        return datasets[0] if len(datasets) == 1 else {'multiple': True, 'count': len(datasets), 'datasets': datasets}


# ========== 图表生成类 ==========

class ChartGenerator:
    """图表生成器"""
    
    @staticmethod
    def limit_data_points(data: List, max_points: int = 50) -> List:
        """限制图表数据点数量"""
        if len(data) <= max_points:
            return data
        step = len(data) // max_points
        return data[::step][:max_points]
    
    @staticmethod
    def create_pie_chart(worksheet, data_range: str, title: str, position: str,
                         labels_range: str = None, width: int = 12, height: int = 8):
        """创建饼图"""
        pie = PieChart()
        pie.title = title
        pie.width = width
        pie.height = height
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        worksheet.add_chart(pie, position)
        return pie
    
    @staticmethod
    def create_bar_chart(worksheet, data_range: str, title: str, position: str,
                         cats_range: str = None, width: int = 12, height: int = 8,
                         y_max: int = None):
        """创建柱状图"""
        bar = BarChart()
        bar.title = title
        bar.type = "col"
        bar.style = 10
        bar.width = width
        bar.height = height
        worksheet.add_chart(bar, position)
        return bar
    
    @staticmethod
    def create_line_chart(worksheet, data_range: str, title: str, position: str,
                          cats_range: str = None, width: int = 12, height: int = 8):
        """创建折线图"""
        line = LineChart()
        line.title = title
        line.style = 10
        line.width = width
        line.height = height
        worksheet.add_chart(line, position)
        return line


# ========== 工作表生成类 ==========

class WorksheetGenerator:
    """工作表生成器"""
    
    def __init__(self, workbook: Workbook):
        self.wb = workbook
    
    def create_student_statistics_sheet(self, student_stats: List[Dict]) -> None:
        """创建按学生统计工作表"""
        ws = self.wb.create_sheet("按学生统计")
        
        col_widths = [15, 12, 10, 10, 10, 10, 12, 12, 12, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        headers = ['学生姓名', '学生ID', '作业数', '总题数', '正确数', '错误数', 
                   '准确率', '选择题准确率', '填空题准确率', '主观题准确率']
        for col, header in enumerate(headers, 1):
            StyleConfig.apply_header_style(ws.cell(row=1, column=col, value=header))
        
        for row_idx, stats in enumerate(student_stats, 2):
            ws.cell(row=row_idx, column=1, value=stats.get('student_name', ''))
            ws.cell(row=row_idx, column=2, value=stats.get('student_id', ''))
            ws.cell(row=row_idx, column=3, value=stats.get('homework_count', 0))
            ws.cell(row=row_idx, column=4, value=stats.get('total_questions', 0))
            ws.cell(row=row_idx, column=5, value=stats.get('correct_count', 0))
            ws.cell(row=row_idx, column=6, value=stats.get('error_count', 0))
            
            acc = stats.get('accuracy', 0)
            acc_cell = ws.cell(row=row_idx, column=7, value=f"{acc * 100:.1f}%")
            acc_cell.fill = StyleConfig.get_accuracy_fill(acc)
            
            ws.cell(row=row_idx, column=8, value=f"{stats.get('choice_accuracy', 0) * 100:.1f}%" if stats.get('choice_total', 0) > 0 else '-')
            ws.cell(row=row_idx, column=9, value=f"{stats.get('fill_accuracy', 0) * 100:.1f}%" if stats.get('fill_total', 0) > 0 else '-')
            ws.cell(row=row_idx, column=10, value=f"{stats.get('subjective_accuracy', 0) * 100:.1f}%" if stats.get('subjective_total', 0) > 0 else '-')
            
            for col in range(1, 11):
                StyleConfig.apply_cell_style(ws.cell(row=row_idx, column=col), center=(col > 1))
        
        ws.freeze_panes = 'A2'
        if student_stats:
            ws.auto_filter.ref = f"A1:J{len(student_stats) + 1}"
    
    def create_page_statistics_sheet(self, page_stats: List[Dict]) -> None:
        """创建按页码统计工作表"""
        ws = self.wb.create_sheet("按页码统计")
        
        col_widths = [20, 10, 10, 10, 10, 10, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        headers = ['书本名称', '页码', '作业数', '总题数', '正确数', '错误数', '准确率']
        for col, header in enumerate(headers, 1):
            StyleConfig.apply_header_style(ws.cell(row=1, column=col, value=header))
        
        for row_idx, stats in enumerate(page_stats, 2):
            ws.cell(row=row_idx, column=1, value=stats.get('book_name', ''))
            ws.cell(row=row_idx, column=2, value=stats.get('page_num', ''))
            ws.cell(row=row_idx, column=3, value=stats.get('homework_count', 0))
            ws.cell(row=row_idx, column=4, value=stats.get('total_questions', 0))
            ws.cell(row=row_idx, column=5, value=stats.get('correct_count', 0))
            ws.cell(row=row_idx, column=6, value=stats.get('error_count', 0))
            
            acc = stats.get('accuracy', 0)
            acc_cell = ws.cell(row=row_idx, column=7, value=f"{acc * 100:.1f}%")
            acc_cell.fill = StyleConfig.get_accuracy_fill(acc)
            
            for col in range(1, 8):
                StyleConfig.apply_cell_style(ws.cell(row=row_idx, column=col), center=(col > 1))
        
        ws.freeze_panes = 'A2'
        if page_stats:
            ws.auto_filter.ref = f"A1:G{len(page_stats) + 1}"

    
    def create_type_statistics_sheet(self, type_stats: List[Dict]) -> None:
        """创建按题型统计工作表"""
        ws = self.wb.create_sheet("按题型统计")
        
        col_widths = [15, 10, 10, 10, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        headers = ['题型', '总题数', '正确数', '错误数', '准确率']
        for col, header in enumerate(headers, 1):
            StyleConfig.apply_header_style(ws.cell(row=1, column=col, value=header))
        
        for row_idx, stats in enumerate(type_stats, 2):
            ws.cell(row=row_idx, column=1, value=stats.get('type_name', ''))
            ws.cell(row=row_idx, column=2, value=stats.get('total', 0))
            ws.cell(row=row_idx, column=3, value=stats.get('correct', 0))
            ws.cell(row=row_idx, column=4, value=stats.get('error_count', 0))
            
            acc = stats.get('accuracy', 0)
            acc_cell = ws.cell(row=row_idx, column=5, value=f"{acc * 100:.1f}%" if stats.get('total', 0) > 0 else '-')
            if stats.get('total', 0) > 0:
                acc_cell.fill = StyleConfig.get_accuracy_fill(acc)
            
            for col in range(1, 6):
                StyleConfig.apply_cell_style(ws.cell(row=row_idx, column=col), center=(col > 1))
        
        ws.freeze_panes = 'A2'
    
    def create_essay_scores_sheet(self, essay_data: Dict) -> None:
        """创建英语作文评分工作表"""
        if not essay_data.get('has_essay'):
            return
        
        ws = self.wb.create_sheet("英语作文评分")
        
        col_widths = [15, 12, 10, 10, 40, 40]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 统计信息
        stats = essay_data.get('stats', {})
        ws['A1'] = "作文评分统计"
        ws['A1'].font = StyleConfig.FONT_SUBTITLE
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"作文数: {stats.get('count', 0)}"
        ws['B2'] = f"平均分: {stats.get('avg_score', 0)}"
        ws['C2'] = f"最高分: {stats.get('max_score', 0)}"
        ws['D2'] = f"最低分: {stats.get('min_score', 0)}"
        
        # 表头
        headers = ['学生姓名', '学生ID', '题号', '参考得分', '综合评价', '改进建议']
        for col, header in enumerate(headers, 1):
            StyleConfig.apply_header_style(ws.cell(row=4, column=col, value=header))
        
        # 数据行
        essays = essay_data.get('essays', [])
        for row_idx, essay in enumerate(essays, 5):
            ws.cell(row=row_idx, column=1, value=essay.get('student_name', ''))
            ws.cell(row=row_idx, column=2, value=essay.get('student_id', ''))
            ws.cell(row=row_idx, column=3, value=essay.get('index', ''))
            ws.cell(row=row_idx, column=4, value=essay.get('score', 0))
            ws.cell(row=row_idx, column=5, value=essay.get('evaluation', '')[:200])
            ws.cell(row=row_idx, column=6, value=essay.get('suggestions', '')[:200])
            
            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                StyleConfig.apply_cell_style(cell, center=(col <= 4))
                if col >= 5:
                    cell.alignment = StyleConfig.ALIGN_LEFT_WRAP
        
        ws.freeze_panes = 'A5'
        if essays:
            ws.auto_filter.ref = f"A4:F{len(essays) + 4}"

    
    def create_enhanced_error_details_sheet(self, error_details: List[Dict]) -> None:
        """创建增强版错误详情工作表"""
        ws = self.wb.create_sheet("错误详情")
        
        col_widths = [10, 15, 8, 8, 12, 15, 15, 15, 10, 10, 15, 10, 10, 40]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        headers = ['序号', '书本', '页码', '题号', '学生姓名', '错误类型',
                   '基准用户答案', 'AI识别答案', '基准判断', 'AI判断',
                   '标准答案', '相似度', '严重程度', '详细说明']
        for col, header in enumerate(headers, 1):
            StyleConfig.apply_header_style(ws.cell(row=1, column=col, value=header))
        
        for row_idx, err in enumerate(error_details, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=err.get('book_name', ''))
            ws.cell(row=row_idx, column=3, value=err.get('page_num', ''))
            ws.cell(row=row_idx, column=4, value=err.get('index', ''))
            ws.cell(row=row_idx, column=5, value=err.get('student_name', ''))
            ws.cell(row=row_idx, column=6, value=err.get('error_type', ''))
            ws.cell(row=row_idx, column=7, value=err.get('base_user_answer', ''))
            ws.cell(row=row_idx, column=8, value=err.get('ai_user_answer', ''))
            ws.cell(row=row_idx, column=9, value=err.get('base_correct', ''))
            ws.cell(row=row_idx, column=10, value=err.get('ai_correct', ''))
            ws.cell(row=row_idx, column=11, value=err.get('standard_answer', ''))
            
            similarity = err.get('similarity')
            ws.cell(row=row_idx, column=12, value=f"{similarity * 100:.1f}%" if similarity is not None else '-')
            
            ws.cell(row=row_idx, column=13, value=err.get('severity', ''))
            ws.cell(row=row_idx, column=14, value=err.get('explanation', '')[:200])
            
            for col in range(1, 15):
                cell = ws.cell(row=row_idx, column=col)
                StyleConfig.apply_cell_style(cell, center=(col <= 13))
                if col == 14:
                    cell.alignment = StyleConfig.ALIGN_LEFT_WRAP
        
        ws.freeze_panes = 'A2'
        if error_details:
            ws.auto_filter.ref = f"A1:N{len(error_details) + 1}"


# ========== 辅助工作表创建函数 ==========

def _create_summary_sheet(ws, task_data: Dict, overall: Dict) -> None:
    """创建评估总结表"""
    # 设置列宽
    for col, width in enumerate([25, 20, 20, 20, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'] = "AI批改效果评估报告"
    ws['A1'].font = StyleConfig.FONT_TITLE
    ws['A1'].alignment = StyleConfig.ALIGN_CENTER
    ws.row_dimensions[1].height = 30
    
    # 导出元数据（右上角）
    ws['E2'] = f"导出版本: 2.0"
    ws['E2'].font = Font(size=9, color="86868B")
    ws['E2'].alignment = Alignment(horizontal='right')
    ws['E3'] = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['E3'].font = Font(size=9, color="86868B")
    ws['E3'].alignment = Alignment(horizontal='right')
    
    # 基本信息
    info_data = [
        ('任务名称', task_data.get('name', '')),
        ('创建时间', task_data.get('created_at', '')),
        ('评估状态', '已完成' if task_data.get('status') == 'completed' else task_data.get('status', '')),
        ('测试条件', task_data.get('test_condition_name', '-')),
    ]
    
    for row_idx, (label, value) in enumerate(info_data, 3):
        ws.cell(row=row_idx, column=1, value=label).font = StyleConfig.FONT_SECTION
        ws.cell(row=row_idx, column=2, value=value)
    
    # 数据集信息
    dataset_info = DataExtractor.extract_dataset_info(task_data)
    if dataset_info:
        row = 8
        ws.cell(row=row, column=1, value="数据集信息").font = StyleConfig.FONT_SUBTITLE
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        if dataset_info.get('multiple'):
            ws.cell(row=row, column=1, value=f"使用了 {dataset_info['count']} 个数据集")
        else:
            ws.cell(row=row, column=1, value="数据集名称")
            ws.cell(row=row, column=2, value=dataset_info.get('name', ''))
            row += 1
            ws.cell(row=row, column=1, value="数据集描述")
            ws.cell(row=row, column=2, value=dataset_info.get('description', ''))
            row += 1
            ws.cell(row=row, column=1, value="创建时间")
            ws.cell(row=row, column=2, value=dataset_info.get('created_at', ''))
            row += 1
            ws.cell(row=row, column=1, value="包含页码")
            ws.cell(row=row, column=2, value=', '.join(map(str, dataset_info.get('pages', []))))
            row += 1
            ws.cell(row=row, column=1, value="题目总数")
            ws.cell(row=row, column=2, value=dataset_info.get('question_count', 0))
        row += 2
    else:
        row = 8
    
    # 核心指标
    ws.cell(row=row, column=1, value="核心评估指标").font = StyleConfig.FONT_SUBTITLE
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    headers = ['指标', '数值', '说明']
    for col, h in enumerate(headers, 1):
        StyleConfig.apply_header_style(ws.cell(row=row, column=col, value=h))
    row += 1
    
    accuracy = overall.get('overall_accuracy', 0)
    total_q = overall.get('total_questions', 0)
    correct_q = overall.get('correct_questions', 0)
    
    metrics = [
        ('总体准确率', f"{accuracy * 100:.1f}%", '正确题目数/总题目数'),
        ('总作业数', overall.get('total_homework', 0), '参与评估的作业数量'),
        ('总题目数', total_q, '所有作业的题目总数'),
        ('正确题目数', correct_q, 'AI批改与基准一致的题目'),
        ('错误题目数', total_q - correct_q, 'AI批改与基准不一致的题目'),
    ]
    
    for label, value, desc in metrics:
        ws.cell(row=row, column=1, value=label).border = StyleConfig.BORDER_THIN
        cell = ws.cell(row=row, column=2, value=value)
        cell.border = StyleConfig.BORDER_THIN
        cell.alignment = StyleConfig.ALIGN_CENTER
        if label == '总体准确率':
            cell.fill = StyleConfig.get_accuracy_fill(accuracy)
        ws.cell(row=row, column=3, value=desc).border = StyleConfig.BORDER_THIN
        row += 1
    
    row += 1
    
    # 题型分类统计
    ws.cell(row=row, column=1, value="题型分类统计").font = StyleConfig.FONT_SUBTITLE
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    type_headers = ['题型', '总数', '正确数', '错误数', '准确率']
    for col, h in enumerate(type_headers, 1):
        StyleConfig.apply_header_style(ws.cell(row=row, column=col, value=h))
    row += 1
    
    by_type = overall.get('by_question_type', {})
    for type_key, type_name in [('choice', '选择题'), ('objective_fill', '客观填空题'), ('subjective', '主观题')]:
        stats = by_type.get(type_key, {})
        total = stats.get('total', 0)
        correct = stats.get('correct', 0)
        acc = stats.get('accuracy', 0)
        
        ws.cell(row=row, column=1, value=type_name).border = StyleConfig.BORDER_THIN
        ws.cell(row=row, column=2, value=total).border = StyleConfig.BORDER_THIN
        ws.cell(row=row, column=3, value=correct).border = StyleConfig.BORDER_THIN
        ws.cell(row=row, column=4, value=total - correct).border = StyleConfig.BORDER_THIN
        acc_cell = ws.cell(row=row, column=5, value=f"{acc * 100:.1f}%" if total > 0 else '-')
        acc_cell.border = StyleConfig.BORDER_THIN
        if total > 0:
            acc_cell.fill = StyleConfig.get_accuracy_fill(acc)
        row += 1
    
    ws.freeze_panes = 'A2'


def _create_error_analysis_sheet(ws, homework_items: List[Dict]) -> None:
    """创建错误分析表"""
    for col, width in enumerate([25, 15, 15, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 统计错误类型分布
    error_distribution = {}
    question_error_count = {}
    
    for item in homework_items:
        if item.get('status') != 'completed':
            continue
        evaluation = item.get('evaluation', {})
        errors = evaluation.get('errors', [])
        page_num = str(item.get('page_num', '?'))
        
        for err in errors:
            err_type = err.get('error_type', '未分类')
            error_distribution[err_type] = error_distribution.get(err_type, 0) + 1
            
            q_idx = err.get('index', '?')
            q_key = f"P{page_num}-{q_idx}"
            question_error_count[q_key] = question_error_count.get(q_key, 0) + 1
    
    # 错误类型分布表
    ws['A1'] = "错误类型分布"
    ws['A1'].font = StyleConfig.FONT_SUBTITLE
    ws.merge_cells('A1:D1')
    
    headers = ['错误类型', '数量', '占比', '说明']
    for col, h in enumerate(headers, 1):
        StyleConfig.apply_header_style(ws.cell(row=2, column=col, value=h))
    
    total_errors = sum(error_distribution.values()) or 1
    error_type_desc = {
        '识别错误-判断正确': '用户答案识别不准确，但最终判断结果正确',
        '识别错误-判断错误': '用户答案识别不准确，且判断结果错误',
        '识别正确-判断错误': '用户答案识别正确，但判断结果错误',
        '缺失题目': 'AI批改结果中缺少该题目',
        'AI识别幻觉': 'AI将错误答案识别为标准答案',
        '格式差异': '答案格式不一致但内容相同'
    }
    
    sorted_errors = sorted(error_distribution.items(), key=lambda x: -x[1])
    for row_idx, (err_type, count) in enumerate(sorted_errors, 3):
        ws.cell(row=row_idx, column=1, value=err_type).border = StyleConfig.BORDER_THIN
        ws.cell(row=row_idx, column=2, value=count).border = StyleConfig.BORDER_THIN
        ws.cell(row=row_idx, column=3, value=f"{count/total_errors*100:.1f}%").border = StyleConfig.BORDER_THIN
        ws.cell(row=row_idx, column=4, value=error_type_desc.get(err_type, '')).border = StyleConfig.BORDER_THIN
    
    # 图表数据区域
    chart_start_row = len(sorted_errors) + 5
    ws.cell(row=chart_start_row, column=1, value="图表数据区").font = StyleConfig.FONT_SECTION
    
    for row_idx, (err_type, count) in enumerate(sorted_errors, chart_start_row + 1):
        ws.cell(row=row_idx, column=1, value=err_type)
        ws.cell(row=row_idx, column=2, value=count)
    
    # 创建饼图
    if sorted_errors:
        pie = PieChart()
        pie.title = "错误类型分布"
        labels = Reference(ws, min_col=1, min_row=chart_start_row+1, max_row=chart_start_row+len(sorted_errors))
        data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_start_row+len(sorted_errors))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 15
        pie.height = 10
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        ws.add_chart(pie, "F2")
    
    # 高频错误题目 TOP10
    top_row = chart_start_row + len(sorted_errors) + 3
    ws.cell(row=top_row, column=1, value="高频错误题目 TOP10").font = StyleConfig.FONT_SUBTITLE
    ws.merge_cells(f'A{top_row}:D{top_row}')
    
    top_headers = ['题目', '错误次数', '占比']
    for col, h in enumerate(top_headers, 1):
        StyleConfig.apply_header_style(ws.cell(row=top_row+1, column=col, value=h))
    
    sorted_questions = sorted(question_error_count.items(), key=lambda x: -x[1])[:10]
    for row_idx, (q_key, count) in enumerate(sorted_questions, top_row+2):
        ws.cell(row=row_idx, column=1, value=q_key).border = StyleConfig.BORDER_THIN
        ws.cell(row=row_idx, column=2, value=count).border = StyleConfig.BORDER_THIN
        ws.cell(row=row_idx, column=3, value=f"{count/total_errors*100:.1f}%").border = StyleConfig.BORDER_THIN
    
    # 柱状图
    if sorted_questions:
        bar = BarChart()
        bar.title = "高频错误题目 TOP10"
        bar.type = "col"
        bar.style = 10
        bar.width = 15
        bar.height = 10
        data = Reference(ws, min_col=2, min_row=top_row+1, max_row=top_row+1+len(sorted_questions))
        cats = Reference(ws, min_col=1, min_row=top_row+2, max_row=top_row+1+len(sorted_questions))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "F18")
    
    ws.freeze_panes = 'A3'


def _create_charts_sheet(ws, overall: Dict) -> None:
    """创建可视化图表表"""
    for col, width in enumerate([20, 15, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 题型准确率对比数据
    ws['A1'] = "题型准确率对比"
    ws['A1'].font = StyleConfig.FONT_SUBTITLE
    
    headers = ['题型', '准确率']
    for col, h in enumerate(headers, 1):
        StyleConfig.apply_header_style(ws.cell(row=2, column=col, value=h))
    
    by_type = overall.get('by_question_type', {})
    type_data = [
        ('选择题', by_type.get('choice', {}).get('accuracy', 0)),
        ('客观填空题', by_type.get('objective_fill', {}).get('accuracy', 0)),
        ('主观题', by_type.get('subjective', {}).get('accuracy', 0))
    ]
    
    for row_idx, (type_name, acc) in enumerate(type_data, 3):
        ws.cell(row=row_idx, column=1, value=type_name)
        ws.cell(row=row_idx, column=2, value=acc * 100)
    
    # 创建柱状图
    if any(acc > 0 for _, acc in type_data):
        bar = BarChart()
        bar.title = "题型准确率对比"
        bar.type = "col"
        bar.style = 10
        bar.width = 12
        bar.height = 8
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        cats = Reference(ws, min_col=1, min_row=3, max_row=5)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "D2")
    
    # 书本准确率对比
    by_book = overall.get('by_book', {})
    if by_book:
        row = 8
        ws.cell(row=row, column=1, value="书本准确率对比").font = StyleConfig.FONT_SUBTITLE
        row += 1
        
        headers = ['书本', '准确率']
        for col, h in enumerate(headers, 1):
            StyleConfig.apply_header_style(ws.cell(row=row, column=col, value=h))
        row += 1
        
        for book_name, stats in list(by_book.items())[:10]:
            ws.cell(row=row, column=1, value=book_name[:20])
            ws.cell(row=row, column=2, value=stats.get('accuracy', 0) * 100)
            row += 1


def _create_homework_details_sheet(ws, homework_items: List[Dict]) -> None:
    """创建作业明细表"""
    col_widths = [12, 20, 8, 15, 12, 10, 10, 10, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    headers = ['作业ID', '书本名称', '页码', '学生姓名', '学生ID', '总题数', 
               '正确数', '错误数', '准确率', '数据集名称']
    for col, header in enumerate(headers, 1):
        StyleConfig.apply_header_style(ws.cell(row=1, column=col, value=header))
    
    row_idx = 2
    for item in homework_items:
        if item.get('status') != 'completed':
            continue
        
        evaluation = item.get('evaluation', {})
        ws.cell(row=row_idx, column=1, value=item.get('homework_id', ''))
        ws.cell(row=row_idx, column=2, value=item.get('book_name', ''))
        ws.cell(row=row_idx, column=3, value=item.get('page_num', ''))
        ws.cell(row=row_idx, column=4, value=get_with_fallback(item, ['student_name', 'studentName'], '未知学生'))
        ws.cell(row=row_idx, column=5, value=get_with_fallback(item, ['student_id', 'studentId'], ''))
        ws.cell(row=row_idx, column=6, value=evaluation.get('total_questions', 0))
        ws.cell(row=row_idx, column=7, value=evaluation.get('correct_count', 0))
        ws.cell(row=row_idx, column=8, value=evaluation.get('error_count', 0))
        
        acc = item.get('accuracy', 0)
        acc_cell = ws.cell(row=row_idx, column=9, value=f"{acc * 100:.1f}%")
        acc_cell.fill = StyleConfig.get_accuracy_fill(acc)
        
        ws.cell(row=row_idx, column=10, value=item.get('matched_dataset_name', '-'))
        
        for col in range(1, 11):
            StyleConfig.apply_cell_style(ws.cell(row=row_idx, column=col), center=(col not in [2, 4, 10]))
        
        row_idx += 1
    
    ws.freeze_panes = 'A2'
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:J{row_idx - 1}"


def _create_question_details_sheet(ws, homework_items: List[Dict], task_data: Dict) -> None:
    """创建题目明细表（增强版）"""
    col_widths = [12, 15, 8, 8, 12, 12, 15, 15, 10, 10, 10, 15, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    headers = ['作业ID', '书本名称', '页码', '题号', '题型', '学生姓名',
               '标准答案', '用户答案', 'AI判断', '基准判断', '相似度', '错误类型', '说明']
    for col, header in enumerate(headers, 1):
        StyleConfig.apply_header_style(ws.cell(row=1, column=col, value=header))
    
    row_idx = 2
    for item in homework_items:
        if item.get('status') != 'completed':
            continue
        
        homework_id = item.get('homework_id', '')
        book_name = item.get('book_name', '')
        page_num = item.get('page_num', '')
        student_name = get_with_fallback(item, ['student_name', 'studentName'], '未知学生')
        
        # 获取基准效果
        base_effects = DataExtractor._get_base_effects_for_homework(item, task_data)
        base_map = {str(q.get('index', '')): q for q in base_effects}
        
        # 获取AI结果
        homework_result = normalize_homework_result(item.get('homework_result', '[]'))
        ai_map = {}
        for q in homework_result:
            ai_map[str(q.get('index', ''))] = q
            for child in q.get('children', []):
                ai_map[str(child.get('index', ''))] = child
        
        # 获取错误信息
        errors = item.get('evaluation', {}).get('errors', [])
        error_map = {str(e.get('index', '')): e for e in errors}
        
        # 遍历所有题目
        all_indices = set(base_map.keys()) | set(ai_map.keys())
        for idx in sorted(all_indices, key=lambda x: safe_int(x, 999)):
            base_item = base_map.get(idx, {})
            ai_item = ai_map.get(idx, {})
            error_info = error_map.get(idx, {})
            
            type_info = classify_question_type(base_item)
            
            ws.cell(row=row_idx, column=1, value=homework_id)
            ws.cell(row=row_idx, column=2, value=book_name)
            ws.cell(row=row_idx, column=3, value=page_num)
            ws.cell(row=row_idx, column=4, value=idx)
            ws.cell(row=row_idx, column=5, value=type_info.get('type_name', '未分类'))
            ws.cell(row=row_idx, column=6, value=student_name)
            ws.cell(row=row_idx, column=7, value=get_with_fallback(base_item, ['answer', 'mainAnswer'], ''))
            ws.cell(row=row_idx, column=8, value=get_with_fallback(ai_item, ['userAnswer', 'user_answer'], ''))
            ws.cell(row=row_idx, column=9, value=get_correct_value(ai_item))
            ws.cell(row=row_idx, column=10, value=get_correct_value(base_item))
            
            similarity = error_info.get('similarity')
            ws.cell(row=row_idx, column=11, value=f"{similarity * 100:.1f}%" if similarity is not None else '-')
            
            ws.cell(row=row_idx, column=12, value=error_info.get('error_type', '-'))
            ws.cell(row=row_idx, column=13, value=error_info.get('explanation', '')[:100])
            
            for col in range(1, 14):
                StyleConfig.apply_cell_style(ws.cell(row=row_idx, column=col), center=(col not in [2, 6, 7, 8, 13]))
            
            row_idx += 1
    
    ws.freeze_panes = 'A2'
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:M{row_idx - 1}"


def _create_ai_report_sheet(ws, overall: Dict) -> None:
    """创建AI分析报告表"""
    for col, width in enumerate([30, 60], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws['A1'] = "AI分析报告"
    ws['A1'].font = StyleConfig.FONT_TITLE
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 30
    
    ai_analysis = overall.get('ai_analysis', {})
    
    row = 3
    sections = [
        ('总体评价', ai_analysis.get('summary', '暂无分析')),
        ('主要问题', ai_analysis.get('main_issues', '暂无分析')),
        ('改进建议', ai_analysis.get('suggestions', '暂无分析')),
        ('详细分析', ai_analysis.get('detailed_analysis', '暂无分析'))
    ]
    
    for title, content in sections:
        ws.cell(row=row, column=1, value=title).font = StyleConfig.FONT_SUBTITLE
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        ws.cell(row=row, column=1, value=content)
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1).alignment = StyleConfig.ALIGN_LEFT_WRAP
        ws.row_dimensions[row].height = 80
        row += 2


# ========== 性能优化工具函数 ==========

def batch_write_rows(worksheet, data: List[List], start_row: int = 1, batch_size: int = 100) -> None:
    """批量写入行数据，减少IO操作"""
    for i in range(0, len(data), batch_size):
        batch = data[i:i+batch_size]
        for row_idx, row_data in enumerate(batch, start_row + i):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)


def stream_write_large_data(worksheet, data_generator, start_row: int = 2) -> int:
    """流式写入大数据，定期清理内存"""
    import gc
    row_idx = start_row
    for row_data in data_generator:
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1
        if row_idx % 1000 == 0:
            gc.collect()
    return row_idx


def lazy_extract_data(task_data: Dict) -> Dict:
    """延迟提取数据，只在需要时计算"""
    homework_items = task_data.get('homework_items', [])
    return {
        'student_stats': lambda: DataExtractor.extract_student_statistics(homework_items),
        'page_stats': lambda: DataExtractor.extract_page_statistics(homework_items),
        'type_stats': lambda: DataExtractor.extract_type_statistics(homework_items),
        'error_details': lambda: DataExtractor.extract_error_details(homework_items, task_data),
        'essay_data': lambda: DataExtractor.extract_essay_scores(homework_items, task_data.get('subject_id', 0))
    }


# ========== 主导出函数 ==========

import logging

logger = logging.getLogger(__name__)


def export_batch_excel_enhanced(task_data: Dict) -> Workbook:
    """
    增强版Excel导出函数
    集成所有新模块：数据提取、样式配置、工作表生成、图表生成
    
    Args:
        task_data: 批量评估任务数据
    
    Returns:
        Workbook: 生成的Excel工作簿
    
    Raises:
        ValueError: 当task_data为空或无效时
        Exception: 当导出过程中发生错误时
    """
    if not task_data:
        raise ValueError("task_data不能为空")
    
    try:
        logger.info(f"开始导出Excel报告，任务ID: {task_data.get('task_id', 'unknown')}")
        
        wb = Workbook()
        generator = WorksheetGenerator(wb)
        
        homework_items = task_data.get('homework_items', [])
        overall = task_data.get('overall_report', {})
        subject_id = task_data.get('subject_id', 0)
        
        # ========== 1. 评估总结表 ==========
        ws_summary = wb.active
        ws_summary.title = "评估总结"
        _create_summary_sheet(ws_summary, task_data, overall)
        
        # ========== 2. 错误分析表 ==========
        ws_error = wb.create_sheet("错误分析")
        _create_error_analysis_sheet(ws_error, homework_items)
        
        # ========== 3. 可视化图表表 ==========
        ws_chart = wb.create_sheet("可视化图表")
        _create_charts_sheet(ws_chart, overall)
        
        # ========== 4. 作业明细表 ==========
        ws_homework = wb.create_sheet("作业明细")
        _create_homework_details_sheet(ws_homework, homework_items)
        
        # ========== 5. 题目明细表 ==========
        ws_questions = wb.create_sheet("题目明细")
        _create_question_details_sheet(ws_questions, homework_items, task_data)
        
        # ========== 6. 错误详情表 (增强版) ==========
        error_details = DataExtractor.extract_error_details(homework_items, task_data)
        generator.create_enhanced_error_details_sheet(error_details)
        
        # ========== 7. AI分析报告表 ==========
        ws_ai = wb.create_sheet("AI分析报告")
        _create_ai_report_sheet(ws_ai, overall)
        
        # ========== 8. 新增统计工作表 ==========
        # 按学生统计
        student_stats = DataExtractor.extract_student_statistics(homework_items)
        generator.create_student_statistics_sheet(student_stats)
        
        # 按页码统计
        page_stats = DataExtractor.extract_page_statistics(homework_items)
        generator.create_page_statistics_sheet(page_stats)
        
        # 按题型统计
        type_stats = DataExtractor.extract_type_statistics(homework_items)
        generator.create_type_statistics_sheet(type_stats)
        
        # 英语作文评分 (条件性创建)
        essay_data = DataExtractor.extract_essay_scores(homework_items, subject_id)
        generator.create_essay_scores_sheet(essay_data)
        
        logger.info(f"Excel报告导出完成，共 {len(homework_items)} 个作业")
        return wb
    
    except Exception as e:
        logger.error(f"Excel导出失败: {str(e)}", exc_info=True)
        raise Exception(f"导出Excel报告失败: {str(e)}")
