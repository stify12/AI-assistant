"""
服务层实现

包含模型服务、相似度服务、Excel服务、存储服务等
"""

import os
import json
import base64
import requests
from typing import List, Dict, Any, Optional, Tuple
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import uuid

from .models import (
    ParsedQuestion, SimilarQuestion, KnowledgePoint,
    DedupeResult, AgentTask, TaskConfig, DifficultyLevel, QuestionType
)


# ============ 图片验证服务 ============

ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


def validate_image_format(filename: str) -> Tuple[bool, str]:
    """
    验证图片格式是否有效
    
    Args:
        filename: 文件名
        
    Returns:
        (是否有效, 错误信息)
    """
    if '.' not in filename:
        return False, "文件名缺少扩展名"
    
    ext = filename.rsplit('.', 1)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return False, f"不支持的图片格式: {ext}，仅支持 JPG/JPEG/PNG"
    
    return True, ""


def validate_image_size(file_size: int) -> Tuple[bool, str]:
    """
    验证图片大小是否在限制内
    
    Args:
        file_size: 文件大小（字节）
        
    Returns:
        (是否有效, 错误信息)
    """
    if file_size > MAX_FILE_SIZE:
        size_mb = file_size / (1024 * 1024)
        return False, f"图片大小 {size_mb:.2f}MB 超过限制 10MB"
    
    return True, ""


# ============ 模型服务 ============

class ModelService:
    """模型服务 - 管理和调用AI模型"""
    
    def __init__(self, config_path: str = "config.json"):
        self.config_path = config_path
        self._load_config()
    
    def _load_config(self):
        """加载配置"""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
        except:
            self.config = {}
    
    def get_available_models(self) -> Dict[str, List[Dict[str, str]]]:
        """获取可用模型列表"""
        return {
            "multimodal": [
                {"name": "doubao-seed-1-8-251228", "provider": "豆包", "desc": "Seed 1.8 多模态模型"},
                {"name": "doubao-seed-1-6-vision-250815", "provider": "豆包", "desc": "Seed 1.6 Vision 视觉模型"},
                {"name": "doubao-1-5-vision-pro-32k-250115", "provider": "豆包", "desc": "Vision Pro 视觉模型"},
            ],
            "text_generation": [
                {"name": "deepseek-v3.2", "provider": "DeepSeek", "desc": "DeepSeek V3.2 文本模型"},
                {"name": "gpt-5-chat-latest", "provider": "OpenAI", "desc": "GPT-5 Chat 文本模型"},
                {"name": "doubao-seed-1-6-251015", "provider": "豆包", "desc": "Seed 1.6 文本模型"},
            ]
        }
    
    def call_multimodal(self, model_name: str, image_base64: str, prompt: str, 
                        timeout: int = 120, max_retries: int = 2) -> str:
        """
        调用多模态模型
        
        Args:
            model_name: 模型名称
            image_base64: Base64编码的图片
            prompt: 提示词
            timeout: 超时时间（秒）
            max_retries: 最大重试次数
            
        Returns:
            模型响应文本
        """
        api_key = self.config.get('api_key', '')
        api_url = self.config.get('api_url', 'https://ark.cn-beijing.volces.com/api/v3/chat/completions')
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": model_name,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}}
                    ]
                }
            ]
        }
        
        # 对于seed-1-8模型，使用minimal模式禁用思考以加快速度
        if 'seed-1-8' in model_name:
            payload["reasoning_effort"] = "minimal"
            payload["max_completion_tokens"] = 65535
        
        last_error = None
        for attempt in range(max_retries + 1):
            try:
                response = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
                response.raise_for_status()
                result = response.json()
                return result.get('choices', [{}])[0].get('message', {}).get('content', '')
            except requests.exceptions.Timeout:
                last_error = f"请求超时（{timeout}秒）"
                if attempt < max_retries:
                    continue
            except requests.exceptions.RequestException as e:
                last_error = str(e)
                if attempt < max_retries:
                    continue
            except Exception as e:
                last_error = str(e)
                break
        
        raise Exception(f"模型调用失败: {last_error}")
    
    def call_text_generation(self, model_name: str, prompt: str) -> str:
        """
        调用文本生成模型
        
        Args:
            model_name: 模型名称
            prompt: 提示词
            
        Returns:
            模型响应文本
        """
        # 根据模型选择不同的API
        if 'deepseek' in model_name.lower():
            return self._call_deepseek(prompt)
        else:
            return self._call_doubao_text(model_name, prompt)
    
    def _call_deepseek(self, prompt: str, timeout: int = 120, max_retries: int = 2) -> str:
        """调用DeepSeek模型"""
        api_key = self.config.get('deepseek_api_key', '')
        api_url = "https://api.deepseek.com/v1/chat/completions"
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": "deepseek-chat",
            "temperature": 0,
            "messages": [{"role": "user", "content": prompt}]
        }
        
        last_error = None
        for attempt in range(max_retries + 1):
            try:
                response = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
                response.raise_for_status()
                result = response.json()
                return result.get('choices', [{}])[0].get('message', {}).get('content', '')
            except requests.exceptions.Timeout:
                last_error = f"请求超时（{timeout}秒）"
                if attempt < max_retries:
                    continue
            except requests.exceptions.RequestException as e:
                last_error = str(e)
                if attempt < max_retries:
                    continue
            except Exception as e:
                last_error = str(e)
                break
        
        raise Exception(f"DeepSeek调用失败: {last_error}")
    
    def _call_doubao_text(self, model_name: str, prompt: str, 
                          timeout: int = 120, max_retries: int = 2) -> str:
        """调用豆包文本模型"""
        api_key = self.config.get('api_key', '')
        api_url = self.config.get('api_url', 'https://ark.cn-beijing.volces.com/api/v3/chat/completions')
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": model_name,
            "messages": [{"role": "user", "content": prompt}]
        }
        
        last_error = None
        for attempt in range(max_retries + 1):
            try:
                response = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
                response.raise_for_status()
                result = response.json()
                return result.get('choices', [{}])[0].get('message', {}).get('content', '')
            except requests.exceptions.Timeout:
                last_error = f"请求超时（{timeout}秒）"
                if attempt < max_retries:
                    continue
            except requests.exceptions.RequestException as e:
                last_error = str(e)
                if attempt < max_retries:
                    continue
            except Exception as e:
                last_error = str(e)
                break
        
        raise Exception(f"豆包模型调用失败: {last_error}")

    # ============ 流式调用方法 ============
    
    def call_multimodal_stream(self, model_name: str, image_base64: str, prompt: str,
                                timeout: int = 180):
        """
        流式调用多模态模型
        
        Args:
            model_name: 模型名称
            image_base64: Base64编码的图片
            prompt: 提示词
            timeout: 超时时间
            
        Yields:
            流式输出的文本片段
        """
        api_key = self.config.get('api_key', '')
        api_url = self.config.get('api_url', 'https://ark.cn-beijing.volces.com/api/v3/chat/completions')
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": model_name,
            "stream": True,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}}
                    ]
                }
            ]
        }
        
        # 对于seed-1-8模型，使用minimal模式禁用思考以加快速度
        if 'seed-1-8' in model_name:
            payload["reasoning_effort"] = "minimal"
            payload["max_completion_tokens"] = 65535
        
        try:
            response = requests.post(api_url, headers=headers, json=payload, 
                                    timeout=timeout, stream=True)
            response.raise_for_status()
            
            usage_info = None
            for line in response.iter_lines():
                if line:
                    line = line.decode('utf-8')
                    if line.startswith('data: '):
                        data = line[6:]
                        if data == '[DONE]':
                            break
                        try:
                            chunk = json.loads(data)
                            if 'usage' in chunk and chunk['usage']:
                                usage_info = chunk['usage']
                            content = chunk.get('choices', [{}])[0].get('delta', {}).get('content', '')
                            if content:
                                yield {'type': 'content', 'content': content}
                        except json.JSONDecodeError:
                            continue
            
            if usage_info:
                yield {'type': 'usage', 'usage': usage_info}
        except Exception as e:
            yield {'type': 'error', 'error': str(e)}
    
    def call_text_stream(self, model_name: str, prompt: str, timeout: int = 180):
        """
        流式调用文本生成模型
        
        Args:
            model_name: 模型名称
            prompt: 提示词
            timeout: 超时时间
            
        Yields:
            流式输出的文本片段
        """
        if 'deepseek' in model_name.lower():
            yield from self._call_deepseek_stream(prompt, timeout)
        else:
            yield from self._call_doubao_text_stream(model_name, prompt, timeout)
    
    def _call_deepseek_stream(self, prompt: str, timeout: int = 180):
        """流式调用DeepSeek模型，最后返回token使用量"""
        api_key = self.config.get('deepseek_api_key', '')
        api_url = "https://api.deepseek.com/v1/chat/completions"
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": "deepseek-chat",
            "temperature": 0,
            "stream": True,
            "stream_options": {"include_usage": True},
            "messages": [{"role": "user", "content": prompt}]
        }
        
        try:
            response = requests.post(api_url, headers=headers, json=payload,
                                    timeout=timeout, stream=True)
            response.raise_for_status()
            
            usage_info = None
            for line in response.iter_lines():
                if line:
                    line = line.decode('utf-8')
                    if line.startswith('data: '):
                        data = line[6:]
                        if data == '[DONE]':
                            break
                        try:
                            chunk = json.loads(data)
                            # 检查是否有usage信息
                            if 'usage' in chunk and chunk['usage']:
                                usage_info = chunk['usage']
                            content = chunk.get('choices', [{}])[0].get('delta', {}).get('content', '')
                            if content:
                                yield {'type': 'content', 'content': content}
                        except json.JSONDecodeError:
                            continue
            
            # 最后返回usage信息
            if usage_info:
                yield {'type': 'usage', 'usage': usage_info}
        except Exception as e:
            yield {'type': 'error', 'error': str(e)}
    
    def _call_doubao_text_stream(self, model_name: str, prompt: str, timeout: int = 180):
        """流式调用豆包文本模型，最后返回token使用量"""
        api_key = self.config.get('api_key', '')
        api_url = self.config.get('api_url', 'https://ark.cn-beijing.volces.com/api/v3/chat/completions')
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        payload = {
            "model": model_name,
            "stream": True,
            "stream_options": {"include_usage": True},
            "messages": [{"role": "user", "content": prompt}]
        }
        
        try:
            response = requests.post(api_url, headers=headers, json=payload,
                                    timeout=timeout, stream=True)
            response.raise_for_status()
            
            usage_info = None
            for line in response.iter_lines():
                if line:
                    line = line.decode('utf-8')
                    if line.startswith('data: '):
                        data = line[6:]
                        if data == '[DONE]':
                            break
                        try:
                            chunk = json.loads(data)
                            if 'usage' in chunk and chunk['usage']:
                                usage_info = chunk['usage']
                            content = chunk.get('choices', [{}])[0].get('delta', {}).get('content', '')
                            if content:
                                yield {'type': 'content', 'content': content}
                        except json.JSONDecodeError:
                            continue
            
            if usage_info:
                yield {'type': 'usage', 'usage': usage_info}
        except Exception as e:
            yield {'type': 'error', 'error': str(e)}


# ============ 相似度服务 ============

class SimilarityService:
    """相似度服务 - 计算文本相似度和知识点去重"""
    
    def calculate_similarity(self, text1: str, text2: str) -> float:
        """
        计算两个文本的相似度
        
        使用SequenceMatcher进行简单的字符串相似度计算
        
        Args:
            text1: 第一个文本
            text2: 第二个文本
            
        Returns:
            相似度分数 (0.0 - 1.0)
        """
        if not text1 or not text2:
            return 0.0
        return SequenceMatcher(None, text1, text2).ratio()
    
    def find_duplicates(self, knowledge_points: List[str], 
                        threshold: float = 0.85) -> List[DedupeResult]:
        """
        查找重复的知识点
        
        Args:
            knowledge_points: 知识点列表
            threshold: 相似度阈值
            
        Returns:
            去重结果列表
        """
        results = []
        processed = set()
        
        for i, point1 in enumerate(knowledge_points):
            if i in processed:
                continue
            
            merged_point = point1
            is_merged = False
            max_similarity = 0.0
            
            for j, point2 in enumerate(knowledge_points):
                if i >= j or j in processed:
                    continue
                
                similarity = self.calculate_similarity(point1, point2)
                
                if similarity >= threshold:
                    processed.add(j)
                    is_merged = True
                    max_similarity = max(max_similarity, similarity)
                    # 保留较短的知识点作为合并后的结果
                    if len(point2) < len(merged_point):
                        merged_point = point2
            
            results.append(DedupeResult(
                original_point=point1,
                merged_point=merged_point,
                similarity_score=max_similarity if is_merged else 1.0,
                is_merged=is_merged
            ))
        
        return results
    
    def dedupe_knowledge_points(self, questions: List[ParsedQuestion],
                                 threshold: float = 0.85) -> Tuple[List[KnowledgePoint], List[DedupeResult]]:
        """
        对题目中的知识点进行去重
        
        Args:
            questions: 解析后的题目列表
            threshold: 相似度阈值
            
        Returns:
            (去重后的知识点列表, 去重结果列表)
        """
        # 收集所有知识点
        all_points = []
        point_map = {}  # 知识点文本 -> KnowledgePoint对象
        
        for question in questions:
            for kp in question.knowledge_points:
                if kp.primary not in point_map:
                    all_points.append(kp.primary)
                    point_map[kp.primary] = kp
        
        # 查找重复
        dedupe_results = self.find_duplicates(all_points, threshold)
        
        # 构建去重后的知识点列表
        unique_points = []
        seen = set()
        
        for result in dedupe_results:
            if result.merged_point not in seen:
                seen.add(result.merged_point)
                if result.merged_point in point_map:
                    unique_points.append(point_map[result.merged_point])
        
        return unique_points, dedupe_results


# ============ Excel服务 ============

class ExcelService:
    """Excel导出服务"""
    
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)
    
    def _create_workbook(self) -> Workbook:
        """创建工作簿"""
        wb = Workbook()
        return wb
    
    def _sanitize_for_excel(self, text: str) -> str:
        """
        清理文本中的特殊字符，使其能在Excel中正常显示
        
        Args:
            text: 原始文本
            
        Returns:
            清理后的文本
        """
        import re
        if not text:
            return text
        
        # 移除 LaTeX 数学模式标记
        text = re.sub(r'\\\(|\\\)', '', text)  # 移除 \( 和 \)
        text = re.sub(r'\\\[|\\\]', '', text)  # 移除 \[ 和 \]
        text = re.sub(r'\$\$?', '', text)      # 移除 $ 和 $$
        
        # 转换常见 LaTeX 命令为可读文本
        text = re.sub(r'\\frac\{([^}]*)\}\{([^}]*)\}', r'(\1)/(\2)', text)  # \frac{a}{b} -> (a)/(b)
        text = re.sub(r'\\sqrt\{([^}]*)\}', r'√(\1)', text)  # \sqrt{x} -> √(x)
        text = re.sub(r'\\times', '×', text)
        text = re.sub(r'\\div', '÷', text)
        text = re.sub(r'\\pm', '±', text)
        text = re.sub(r'\\leq', '≤', text)
        text = re.sub(r'\\geq', '≥', text)
        text = re.sub(r'\\neq', '≠', text)
        text = re.sub(r'\\cdot', '·', text)
        text = re.sub(r'\\ldots', '...', text)
        text = re.sub(r'\\left|\\right', '', text)  # 移除 \left 和 \right
        text = re.sub(r'\\[a-zA-Z]+', '', text)  # 移除其他未处理的 LaTeX 命令
        
        # 移除 Excel 不支持的控制字符
        text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
        
        return text
    
    def _style_header(self, ws, headers: List[str]):
        """设置表头样式"""
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def export_parse_result(self, questions: List[ParsedQuestion]) -> str:
        """
        导出解析结果Excel
        
        列：序号、图片来源、题目内容、学科分类、题目类型、难度等级、一级知识点、二级知识点、知识点解析
        """
        from datetime import datetime
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "题目解析结果"
        
        headers = ["序号", "图片来源", "题目内容", "学科分类", "题目类型", 
                   "难度等级", "一级知识点", "二级知识点", "解题思路"]
        self._style_header(ws, headers)
        
        row = 2
        for idx, question in enumerate(questions, 1):
            for kp in question.knowledge_points:
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=question.image_source)
                ws.cell(row=row, column=3, value=self._sanitize_for_excel(question.content))
                ws.cell(row=row, column=4, value=question.subject)
                ws.cell(row=row, column=5, value=question.question_type.value)
                ws.cell(row=row, column=6, value=question.difficulty.value)
                ws.cell(row=row, column=7, value=self._sanitize_for_excel(kp.primary))
                ws.cell(row=row, column=8, value=self._sanitize_for_excel(kp.secondary))
                ws.cell(row=row, column=9, value=self._sanitize_for_excel(kp.analysis))
                row += 1
        
        # 调整列宽
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"题目解析结果_{timestamp}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)
        return filepath
    
    def export_similar_questions(self, questions: List[SimilarQuestion],
                                  knowledge_points: Dict[str, KnowledgePoint]) -> str:
        """
        导出类题结果Excel
        
        列：序号、一级知识点、二级知识点、类题内容、类题答案、解题步骤、难度等级、题目类型
        """
        from datetime import datetime
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "类题练习"
        
        headers = ["序号", "一级知识点", "二级知识点", "题目内容", 
                   "参考答案", "解题步骤", "难度等级", "题目类型"]
        self._style_header(ws, headers)
        
        for idx, sq in enumerate(questions, 1):
            # 优先使用类题自带的知识点，否则从映射中获取
            primary = getattr(sq, 'primary', '') or ''
            secondary = getattr(sq, 'secondary', '') or ''
            if not primary:
                kp = knowledge_points.get(sq.knowledge_point_id)
                if kp:
                    primary = kp.primary
                    secondary = kp.secondary
            
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=self._sanitize_for_excel(primary))
            ws.cell(row=idx+1, column=3, value=self._sanitize_for_excel(secondary))
            ws.cell(row=idx+1, column=4, value=self._sanitize_for_excel(sq.content))
            ws.cell(row=idx+1, column=5, value=self._sanitize_for_excel(sq.answer))
            ws.cell(row=idx+1, column=6, value=self._sanitize_for_excel(sq.solution_steps))
            ws.cell(row=idx+1, column=7, value=sq.difficulty.value)
            ws.cell(row=idx+1, column=8, value=sq.question_type.value)
        
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"类题练习_{timestamp}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)
        return filepath
    
    def export_full_result(self, parsed: List[ParsedQuestion],
                           similar: List[SimilarQuestion]) -> str:
        """
        导出完整结果Excel
        
        列：序号、图片来源、原题内容、学科分类、题目类型、难度等级、
            一级知识点、二级知识点、知识点解析、类题内容、类题答案、解题步骤
        """
        from datetime import datetime
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "完整数据汇总"
        
        headers = ["序号", "图片来源", "原题内容", "学科分类", "题目类型", "难度等级",
                   "一级知识点", "二级知识点", "解题思路", "类题内容", "参考答案", "解题步骤"]
        self._style_header(ws, headers)
        
        # 构建知识点ID到类题的映射
        kp_to_similar = {}
        for sq in similar:
            if sq.knowledge_point_id not in kp_to_similar:
                kp_to_similar[sq.knowledge_point_id] = []
            kp_to_similar[sq.knowledge_point_id].append(sq)
        
        row = 2
        for idx, question in enumerate(parsed, 1):
            for kp in question.knowledge_points:
                similar_qs = kp_to_similar.get(kp.id, [])
                if similar_qs:
                    for sq in similar_qs:
                        ws.cell(row=row, column=1, value=idx)
                        ws.cell(row=row, column=2, value=question.image_source)
                        ws.cell(row=row, column=3, value=self._sanitize_for_excel(question.content))
                        ws.cell(row=row, column=4, value=question.subject)
                        ws.cell(row=row, column=5, value=question.question_type.value)
                        ws.cell(row=row, column=6, value=question.difficulty.value)
                        ws.cell(row=row, column=7, value=self._sanitize_for_excel(kp.primary))
                        ws.cell(row=row, column=8, value=self._sanitize_for_excel(kp.secondary))
                        ws.cell(row=row, column=9, value=self._sanitize_for_excel(kp.analysis))
                        ws.cell(row=row, column=10, value=self._sanitize_for_excel(sq.content))
                        ws.cell(row=row, column=11, value=self._sanitize_for_excel(sq.answer))
                        ws.cell(row=row, column=12, value=self._sanitize_for_excel(sq.solution_steps))
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=question.image_source)
                    ws.cell(row=row, column=3, value=self._sanitize_for_excel(question.content))
                    ws.cell(row=row, column=4, value=question.subject)
                    ws.cell(row=row, column=5, value=question.question_type.value)
                    ws.cell(row=row, column=6, value=question.difficulty.value)
                    ws.cell(row=row, column=7, value=self._sanitize_for_excel(kp.primary))
                    ws.cell(row=row, column=8, value=self._sanitize_for_excel(kp.secondary))
                    ws.cell(row=row, column=9, value=self._sanitize_for_excel(kp.analysis))
                    row += 1
        
        for col in range(1, 13):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"知识点类题完整数据_{timestamp}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)
        return filepath
    
    def export_dedupe_result(self, results: List[DedupeResult]) -> str:
        """
        导出去重结果Excel
        
        列：原知识点、合并后知识点、相似度分数、是否合并
        """
        from datetime import datetime
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "知识点去重"
        
        headers = ["原知识点", "合并后知识点", "相似度", "是否合并"]
        self._style_header(ws, headers)
        
        for idx, result in enumerate(results, 1):
            ws.cell(row=idx+1, column=1, value=result.original_point)
            ws.cell(row=idx+1, column=2, value=result.merged_point)
            ws.cell(row=idx+1, column=3, value=f"{result.similarity_score:.2f}")
            ws.cell(row=idx+1, column=4, value="是" if result.is_merged else "否")
        
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"知识点去重结果_{timestamp}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)
        return filepath


# ============ 存储服务 ============

class StorageService:
    """存储服务 - 管理任务数据的持久化"""
    
    def __init__(self, storage_dir: str = "knowledge_tasks"):
        self.storage_dir = storage_dir
        self.preference_file = os.path.join(storage_dir, "preferences.json")
        os.makedirs(storage_dir, exist_ok=True)
    
    def save_task(self, task: AgentTask) -> str:
        """
        保存任务
        
        Args:
            task: 任务对象
            
        Returns:
            保存的文件路径
        """
        filepath = os.path.join(self.storage_dir, f"{task.task_id}.json")
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(task.to_json())
        return filepath
    
    def load_task(self, task_id: str) -> Optional[AgentTask]:
        """
        加载任务
        
        Args:
            task_id: 任务ID
            
        Returns:
            任务对象，如果不存在则返回None
        """
        filepath = os.path.join(self.storage_dir, f"{task_id}.json")
        if not os.path.exists(filepath):
            return None
        
        with open(filepath, 'r', encoding='utf-8') as f:
            return AgentTask.from_json(f.read())
    
    def list_tasks(self) -> List[str]:
        """
        列出所有任务ID
        
        Returns:
            任务ID列表
        """
        tasks = []
        for filename in os.listdir(self.storage_dir):
            if filename.endswith('.json') and filename != 'preferences.json':
                tasks.append(filename[:-5])
        return tasks
    
    def delete_task(self, task_id: str) -> bool:
        """
        删除任务
        
        Args:
            task_id: 任务ID
            
        Returns:
            是否删除成功
        """
        filepath = os.path.join(self.storage_dir, f"{task_id}.json")
        if os.path.exists(filepath):
            os.remove(filepath)
            return True
        return False
    
    def save_preferences(self, preferences: Dict[str, Any]) -> None:
        """
        保存用户偏好设置
        
        Args:
            preferences: 偏好设置字典
        """
        with open(self.preference_file, 'w', encoding='utf-8') as f:
            json.dump(preferences, f, ensure_ascii=False, indent=2)
    
    def load_preferences(self) -> Dict[str, Any]:
        """
        加载用户偏好设置
        
        Returns:
            偏好设置字典
        """
        if not os.path.exists(self.preference_file):
            return {
                "multimodal_model": "doubao-seed-1-8-251228",
                "text_model": "deepseek-v3.2",
                "similarity_threshold": 0.85,
                "questions_per_point": 1
            }
        
        with open(self.preference_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def update_similar_question(self, task_id: str, question_id: str, 
                                 updates: Dict[str, Any]) -> bool:
        """
        更新类题内容（用户编辑）
        
        Args:
            task_id: 任务ID
            question_id: 类题ID
            updates: 更新内容
            
        Returns:
            是否更新成功
        """
        task = self.load_task(task_id)
        if not task:
            return False
        
        for sq in task.similar_questions:
            if sq.id == question_id:
                if 'content' in updates:
                    sq.content = updates['content']
                if 'answer' in updates:
                    sq.answer = updates['answer']
                if 'solution_steps' in updates:
                    sq.solution_steps = updates['solution_steps']
                self.save_task(task)
                return True
        
        return False
